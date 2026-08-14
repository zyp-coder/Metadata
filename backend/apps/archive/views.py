from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import F, Case, When, Value, IntegerField, Count
from django.utils import timezone
from .models import (
    Archive, ArchiveRecord, ArchiveRecordDetail, ArchiveRecordVersion,
    ArchiveSchemaSnapshot, ArchiveSyncLog, ArchiveOperationLog, ArchiveApi,
    ArchiveChangeBatch, ArchiveChangeDetail, ConsistencyIssue,
    ConsistencyCheckRule, ApiKey, ApiKeyGrant, ApiCallLog,
)
from .serializers import (
    ArchiveListSerializer, ArchiveDetailSerializer, ArchiveCreateSerializer,
    ArchiveRecordListSerializer, ArchiveRecordDetailSerializer,
    ArchiveRecordCreateSerializer, ArchiveRecordUpdateSerializer,
    VersionSerializer, RollbackSerializer, GlobalVersionSerializer,
    SyncLogSerializer, OperationLogSerializer, ArchiveApiSerializer,
    ChangeBatchSerializer, ChangeDetailSerializer, ConsistencyIssueSerializer,
    ConsistencyCheckRuleSerializer, ApiKeySerializer, ApiCallLogSerializer,
)


def _field_released(f, sf):
    """档案字段门控：判断一个物理字段是否最终进入档案。

    统一三分类口径（与建模侧 standard_fields 一致）：
    1. 物理 → 概念：f.release_to_concept 为 True；否则直接不释放。
    2. 概念 → 档案：
       - 有 StandardField（组合）：sf.status='active' 且 sf.is_active 且 sf.release_to_archive
       - 无 StandardField（solo）：f.archive_category='base' 且 f.release_to_archive
    """
    if not getattr(f, 'release_to_concept', True):
        return False
    if sf is not None:
        if getattr(sf, 'status', 'active') != 'active':
            return False
        if not getattr(sf, 'is_active', True):
            return False
        return getattr(sf, 'release_to_archive', True)
    if getattr(f, 'archive_category', '') != 'base':
        return False
    return getattr(f, 'release_to_archive', True)


def _generate_schema_from_domain(domain):
    """从域的所有物理字段生成 schema，StandardField 信息优先，并应用档案字段门控。

    逻辑：
    1. 按建模分组树 DFS 序（父在前、子紧随）确定分组顺序，未分组排最后
    2. 按释放规则（_field_released）过滤未释放到档案的字段
    3. 按输出 code 去重（首次出现为准）
    4. 如果字段有 StandardField，用 StandardField 的 name/type；否则用物理字段自身
    5. 每个字段携带 group_path（根→叶分组名路径）供前端嵌套层级渲染
    6. 并入已释放到档案的计算字段：有分组的随真实分组排序（组内排物理字段之后），
       未分组的兑底到末尾虚拟「计算字段」组
    """
    from apps.modeling.models import Field, StandardField, ComputedField, FieldGroup

    # 分组树 DFS 序（与建模侧左栏树展示顺序一致）+ 完整路径映射
    group_order = {}
    group_paths = {}
    all_groups = list(FieldGroup.objects.filter(domain=domain).order_by('sort_order', 'id'))
    children_map = {}
    for g in all_groups:
        children_map.setdefault(g.parent_id, []).append(g)

    def _walk(parent_id, path):
        for g in children_map.get(parent_id, []):
            group_order[g.id] = len(group_order)
            group_paths[g.id] = path + [g.name]
            _walk(g.id, group_paths[g.id])

    _walk(None, [])

    fields = Field.objects.filter(
        table__domain=domain, status=Field.Status.ACTIVE
    ).select_related('table', 'group', 'standard_field')
    # 排序：分组 DFS 序优先（未分组排最后），组内按 sort_order/id
    fields = sorted(fields, key=lambda f: (
        group_order.get(f.group_id, 10 ** 9), f.sort_order, f.id
    ))

    # 预加载 StandardField 映射
    sf_ids = set()
    for f in fields:
        if f.standard_field_id:
            sf_ids.add(f.standard_field_id)
    sf_map = {}
    if sf_ids:
        for sf in StandardField.objects.filter(id__in=sf_ids):
            sf_map[sf.id] = sf

    entries = []  # (排序键, schema项)：物理/计算字段统一按分组 DFS 序排列
    seen = set()
    for f in fields:
        sf = sf_map.get(f.standard_field_id) if f.standard_field_id else None
        # 两层释放门控：未释放到档案的字段直接跳过
        if not _field_released(f, sf):
            continue
        out_code = sf.standard_code if sf else (f.code or f.name)
        if out_code in seen:
            continue
        seen.add(out_code)

        sort_key = (group_order.get(f.group_id, 10 ** 9), 0, f.sort_order, f.id)
        if sf:
            entries.append((sort_key, {
                'code': sf.standard_code,
                'name': sf.standard_name or f.comment or f.name,
                'type': sf.field_type,
                'note': sf.note,
                'group': f.group.name if f.group else '',
                'group_path': group_paths.get(f.group_id, []),
                'table': f.table.name,
                'distinct_values': f.distinct_values or [],
                'ownership': getattr(sf, 'ownership', 'archive') or 'archive',
            }))
        else:
            entries.append((sort_key, {
                'code': out_code,
                'name': f.comment or f.name,
                'type': f.field_type,
                'group': f.group.name if f.group else '',
                'group_path': group_paths.get(f.group_id, []),
                'table': f.table.name,
                'distinct_values': f.distinct_values or [],
                'ownership': getattr(f, 'ownership', 'archive') or 'archive',
            }))

    # 并入计算字段（已释放到档案 + 启用状态）：有分组用真实分组，未分组兑底「计算字段」虚拟组
    for cf in ComputedField.objects.filter(
        domain=domain, status='active', release_to_archive=True
    ).order_by('execution_order'):
        if cf.code in seen:
            continue
        seen.add(cf.code)
        if cf.group_id and cf.group_id in group_paths:
            grp_path = group_paths[cf.group_id]
            grp_name = grp_path[-1]
            sort_key = (group_order.get(cf.group_id, 10 ** 9), 1, cf.execution_order, cf.id)
        else:
            grp_path = ['计算字段']
            grp_name = '计算字段'
            sort_key = (10 ** 9 + 1, 1, cf.execution_order, cf.id)
        entries.append((sort_key, {
            'code': cf.code,
            'name': cf.name,
            'type': cf.output_type,
            'source': 'computed',
            'group': grp_name,
            'group_path': grp_path,
            'ownership': 'archive',
        }))

    entries.sort(key=lambda e: e[0])
    return [item for _, item in entries]


def _merge_record_data(record, schema):
    """双层合并：source_data 为底 + archive 字段 manual_data 覆盖 + 计算字段保留。

    规则（按 schema 逐字段，另并入 source_data 中 schema 外的遗留键）：
    - source='computed'（计算字段）→ 保留 record.data 现值（重算由 computed_service 负责）；
    - ownership='source' → 取 source_data（manual_data 遗留键一并清除）；
    - ownership='archive'（缺省）→ manual_data 有键取 manual，否则取 source_data。

    同时重建 lineage：manual 命中 → source='manual'（保留原有登记信息），否则 source='sync'。
    返回 (merged_data, lineage)，并可能就地清理 record.manual_data 的非法键。
    """
    source_data = record.source_data or {}
    manual_data = dict(record.manual_data or {})
    old_data = record.data or {}
    old_lineage = record.lineage or {}
    now_iso = timezone.now().isoformat()

    merged = {}
    lineage = {}
    schema_codes = set()
    for item in (schema or []):
        code = item.get('code')
        if not code:
            continue
        schema_codes.add(code)
        if item.get('source') == 'computed':
            if code in old_data:
                merged[code] = old_data[code]
                if code in old_lineage:
                    lineage[code] = old_lineage[code]
            manual_data.pop(code, None)  # 计算字段不允许人工覆盖
            continue
        ownership = item.get('ownership') or 'archive'
        if ownership == 'source':
            manual_data.pop(code, None)  # 源系统维护：清除遗留人工覆盖
            if code in source_data:
                merged[code] = source_data[code]
                lineage[code] = old_lineage.get(code) if old_lineage.get(code, {}).get('source') == 'sync' \
                    else {'source': 'sync', 'updated_at': now_iso}
            continue
        # ownership='archive'
        if code in manual_data:
            merged[code] = manual_data[code]
            lineage[code] = old_lineage.get(code) if old_lineage.get(code, {}).get('source') in ('manual', 'resolve') \
                else {'source': 'manual', 'updated_at': now_iso}
        elif code in source_data:
            merged[code] = source_data[code]
            lineage[code] = old_lineage.get(code) if old_lineage.get(code, {}).get('source') == 'sync' \
                else {'source': 'sync', 'updated_at': now_iso}
    # source_data / manual_data 中 schema 外的遗留键（如字段已从模型移除但历史数据仍在）保留展示
    for code, value in manual_data.items():
        if code not in schema_codes and code not in merged:
            merged[code] = value
            if code in old_lineage:
                lineage[code] = old_lineage[code]
    for code, value in source_data.items():
        if code not in schema_codes and code not in merged:
            merged[code] = value
            if code in old_lineage:
                lineage[code] = old_lineage[code]
    record.manual_data = manual_data
    return merged, lineage


def _validate_primary_fields(domain):
    """校验域内参与档案的组合字段是否都已设置有效主字段。

    返回未设置（或主字段已失效）的组合字段列表，用于刷新前拦截：
    所有组合字段必须有主字段（数据源头），否则拉取时无法确定取数口径。
    """
    from apps.modeling.models import StandardField
    missing = []
    qs = StandardField.objects.filter(
        domain=domain, status='active', is_active=True, release_to_archive=True
    ).prefetch_related('members__table')
    for sf in qs:
        members = [m for m in sf.members.all() if m.status == 'active']
        if not members:
            continue
        if not sf.primary_field_id or sf.primary_field_id not in {m.id for m in members}:
            missing.append({'sf_id': sf.id, 'code': sf.standard_code,
                            'name': sf.standard_name or sf.standard_code})
    return missing


class ArchiveViewSet(viewsets.ModelViewSet):
    """档案配置 API"""
    queryset = Archive.objects.select_related('domain').all()
    filterset_fields = ['domain', 'status']
    search_fields = ['name']

    def get_serializer_class(self):
        if self.action == 'list':
            return ArchiveListSerializer
        elif self.action == 'create':
            return ArchiveCreateSerializer
        return ArchiveDetailSerializer

    def perform_create(self, serializer):
        """创建档案时，自动从域的所有物理字段生成 schema 快照"""
        archive = serializer.save()
        domain = archive.domain
        schema = _generate_schema_from_domain(domain)
        archive.schema = schema
        archive.save(update_fields=['schema'])

        # 记录操作日志
        ArchiveOperationLog.objects.create(
            archive=archive,
            operator=archive.created_by or 'system',
            operation_type=ArchiveOperationLog.OperationType.CREATE,
            change_summary={'schema_field_count': len(schema)},
        )

    def perform_destroy(self, instance):
        """分批删除关联记录，绕过 SQLite 变量数上限（too many SQL variables）。

        策略（按依赖顺序清理，避免 Django 6.0 Collector combined_updates 优化
        将过多 PK 塞入单条 UPDATE/SET NULL 语句，超 SQLite 999 变量上限）：

        1. ArchiveRecordVersion：无反向 FK 指向它，可直接按 record_id 分批 DELETE。
        2. ArchiveRecordDetail：有 ArchiveChangeDetail.detail_group FK （SET_NULL）
           指向它 → 先清空 ArchiveChangeDetail 反向引用，再按 detail PK 分批 DELETE。
        3. ArchiveRecord：无 SET_NULL 反向 FK，直接按 PK 分批 DELETE。
        4. instance.delete() CASCADE 兜底其余量小的关联（sync_logs/apis 等）。
        """
        from itertools import islice
        from django.db import connection

        RECORD_BATCH = 500      # 取 ArchiveRecord PK 的批次大小
        DETAIL_PK_BATCH = 200   # 删除 ArchiveRecordDetail PK 的批次大小

        qs = instance.records.all()
        pk_iter = iter(qs.values_list('pk', flat=True).iterator())

        while True:
            batch = list(islice(pk_iter, RECORD_BATCH))
            if not batch:
                break

            # --- 1. 删版本：无反向 FK → safe ---
            ArchiveRecordVersion.objects.filter(record_id__in=batch).delete()

            # --- 2. 删明细行：先清 ArchiveChangeDetail 反指 FK，再分批删 ---
            # 收集本条记录批次下所有明细 PK
            detail_pks = list(ArchiveRecordDetail.objects.filter(
                record_id__in=batch
            ).values_list('pk', flat=True).iterator())

            for i in range(0, len(detail_pks), DETAIL_PK_BATCH):
                pk_batch = detail_pks[i:i + DETAIL_PK_BATCH]
                # 先清空反指（SET_NULL），消除 collector combined_updates
                ArchiveChangeDetail.objects.filter(
                    detail_group_id__in=pk_batch
                ).update(detail_group=None)
                # 再删明细行
                ArchiveRecordDetail.objects.filter(pk__in=pk_batch).delete()

        # --- 3. 删记录行：无反向 FK → safe ---
        pk_iter = iter(qs.values_list('pk', flat=True).iterator())
        while True:
            batch = list(islice(pk_iter, RECORD_BATCH))
            if not batch:
                break
            ArchiveRecord.objects.filter(pk__in=batch).delete()

        # 4. instance.delete() CASCADE 兜底其余量小的关联
        instance.delete()

    @action(detail=True, methods=['post'], url_path='sync-schema')
    def sync_schema(self, request, pk=None):
        """同步模型变更：将域的最新模型更新到档案 schema，并从数据源拉取实际数据"""
        archive = self.get_object()
        domain = archive.domain
        operated_by = request.data.get('operated_by', 'system')

        # 先备份当前 schema
        old_schema = archive.schema

        # 用统一的 helper 重新生成 schema（包含所有物理字段）
        new_schema = _generate_schema_from_domain(domain)

        # 构建 code→field_type 映射
        schema_type_map = {item['code']: item['type'] for item in new_schema}

        archive.schema = new_schema
        archive.schema_version += 1
        archive.save(update_fields=['schema', 'schema_version', 'updated_at'])

        # ===== 从数据源拉取实际数据 =====
        sync_stats = {'records_created': 0, 'records_updated': 0, 'tables_synced': 0, 'errors': []}
        try:
            sync_stats = self._sync_data_from_sources(archive, domain, schema_type_map, operated_by)
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f'同步数据失败: {e}')
            sync_stats['errors'].append(str(e))

        # ===== 数据拉取完成后，触发计算字段批量重算 =====
        try:
            from apps.modeling.computed_service import batch_recalculate
            recalc_result = batch_recalculate(domain.id)
            sync_stats['computed_recalculated'] = recalc_result
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f'计算字段重算失败: {e}')
            sync_stats['computed_recalculated'] = {'error': str(e)}

        # 记录操作日志
        ArchiveOperationLog.objects.create(
            archive=archive,
            operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.SCHEMA_SYNC,
            change_summary={
                'old_schema_version': archive.schema_version - 1,
                'new_schema_version': archive.schema_version,
                'old_field_count': len(old_schema),
                'new_field_count': len(new_schema),
                'sync_stats': sync_stats,
            },
        )
        result = ArchiveDetailSerializer(archive).data
        result['sync_stats'] = sync_stats
        return Response(result)

    @action(detail=True, methods=['post'], url_path='refresh-data')
    def refresh_data(self, request, pk=None):
        """立即刷新数据：仅从数据源整层刷新 source_data + 重算计算字段。

        不重生成 schema、不 bump schema_version（结构变更走 sync-schema）。
        """
        archive = self.get_object()
        operated_by = request.data.get('operated_by', 'system')
        stats = refresh_archive_data(archive, operated_by)
        result = ArchiveDetailSerializer(archive).data
        result['sync_stats'] = stats
        return Response(result)

    @action(detail=True, methods=['get'], url_path='refresh-preview')
    def refresh_preview(self, request, pk=None):
        """刷新预检（dry-run，零写入）：对比源与档案的 schema 变化 + 试算数据变化。

        供前端「立即刷新」预检弹窗展示；确认后 schema 有变走 sync-schema，无变走 refresh-data。
        """
        archive = self.get_object()
        domain = archive.domain
        old_schema = archive.schema or []
        new_schema = _generate_schema_from_domain(domain) if domain else old_schema

        # ===== schema 差异 =====
        old_map = {i['code']: i for i in old_schema if i.get('code')}
        new_map = {i['code']: i for i in new_schema if i.get('code')}
        added = [{'code': c, 'name': new_map[c].get('name') or c} for c in new_map if c not in old_map]
        removed = [{'code': c, 'name': old_map[c].get('name') or c} for c in old_map if c not in new_map]
        changed = []
        attr_labels = (('name', '名称'), ('type', '类型'), ('ownership', '维护方'), ('group_path', '分组'))
        ownership_labels = {'source': '源系统维护', 'archive': '档案维护'}
        for c, ni in new_map.items():
            oi = old_map.get(c)
            if not oi:
                continue
            diffs = []
            for attr, label in attr_labels:
                ov, nv = oi.get(attr), ni.get(attr)
                if attr == 'ownership':
                    ov, nv = ov or 'archive', nv or 'archive'
                if ov != nv:
                    if attr == 'ownership':
                        ov, nv = ownership_labels.get(ov, ov), ownership_labels.get(nv, nv)
                    diffs.append({'attr': label,
                                  'old': ' / '.join(ov) if isinstance(ov, list) else ov,
                                  'new': ' / '.join(nv) if isinstance(nv, list) else nv})
            if diffs:
                changed.append({'code': c, 'name': ni.get('name') or c, 'changes': diffs})
        schema_changes = {
            'added': added, 'removed': removed, 'changed': changed,
            'has_changes': bool(added or removed or changed),
        }

        # ===== 数据变化试算（用新 schema 口径） =====
        if domain:
            data_changes = self._preview_data_changes(archive, domain, new_schema)
        else:
            data_changes = {'tables_checked': 0, 'would_create': 0, 'would_update': 0,
                            'would_deactivate': 0, 'changes_sample': [], 'errors': ['档案未关联域']}
        data_changes['has_changes'] = bool(
            data_changes.get('would_create') or data_changes.get('would_update')
            or data_changes.get('would_deactivate'))
        return Response({'schema_changes': schema_changes, 'data_changes': data_changes})

    @action(detail=True, methods=['post'], url_path='consistency-check')
    def consistency_check(self, request, pk=None):
        """一致性检查（支持 4 种检查类型）：
        1. composite_member: 组合字段非主字段成员值≠主字段值
        2. archive_source_diff: 档案侧人工覆盖与源侧数据差异
        3. orphan_source_record: 源侧数据无法关联主表主键
        4. schema_drift: 档案 schema 与当前建模结构不一致

        已失效的规则（ConsistencyCheckRule.disabled=True）不产生新差异。
        零写入档案数据、不回写任何源表（Hub 式 MDM 宪法）。
        """
        from apps.modeling.models import Table, Field

        archive = self.get_object()
        domain = archive.domain
        if not domain:
            return Response({'error': '档案未关联域'}, status=status.HTTP_400_BAD_REQUEST)

        # 加载已失效的规则集合: (check_type, field_code, member_source)
        disabled_rules = set()
        for rule in ConsistencyCheckRule.objects.filter(archive=archive, disabled=True):
            disabled_rules.add((rule.check_type, rule.field_code, rule.member_source))

        def _is_rule_disabled(check_type, field_code='', member_source=''):
            return (check_type, field_code, member_source) in disabled_rules

        schema_type_map = {i['code']: i['type'] for i in (archive.schema or []) if i.get('code')}
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in (archive.schema or [])}
        code_to_physical = self._build_code_to_physical(domain, schema_type_map)
        code_checks = self._build_code_checks(domain)
        stats = {'checked_fields': len(code_checks), 'tables_checked': 0,
                 'mismatch_count': 0, 'mismatch_records': 0,
                 'new_issues': 0, 'reopened_issues': 0, 'resolved_issues': 0,
                 'open_total': 0, 'errors': [], 'checked_at': timezone.now().isoformat(),
                 'by_type': {}}

        # 主键字段
        primary_table = domain.get_primary_table()
        pk_fields = []
        if primary_table:
            pk_fields = list(Field.objects.filter(
                table=primary_table, is_primary_key=True, status=Field.Status.ACTIVE
            ).values_list('code', flat=True))
        if not pk_fields:
            first_pk = Field.objects.filter(
                table__domain=domain, is_primary_key=True, status=Field.Status.ACTIVE
            ).first()
            if first_pk:
                pk_fields = [first_pk.code]

        now = timezone.now()
        all_mismatches = []  # 汇总所有检查类型的差异

        # ===== 检查类型 1: composite_member =====
        cm_stats = {'new': 0, 'reopened': 0, 'resolved': 0, 'open': 0}
        if code_checks and pk_fields:
            cc_primary_values, cc_member_values = {}, {}
            for table in Table.objects.filter(domain=domain, status=Table.Status.ACTIVE):
                try:
                    rows = self._query_local_table(table) if not table.data_source \
                        else self._query_external_table(table)
                    if rows is not None:
                        self._collect_check_values(table, rows, code_checks, pk_fields,
                                                   code_to_physical, cc_primary_values, cc_member_values)
                        stats['tables_checked'] += 1
                except Exception as e:
                    stats['errors'].append(f'{table.name}: {str(e)}')
            mismatches = self._collect_full_mismatches(
                code_checks, cc_primary_values, cc_member_values, field_name_map)
            for m in mismatches:
                m['check_type'] = ConsistencyIssue.CheckType.COMPOSITE_MEMBER
            all_mismatches.extend(mismatches)
            cm_stats['open'] = len(mismatches)
        stats['by_type']['composite_member'] = cm_stats

        # ===== 检查类型 2: archive_source_diff =====
        asd_stats = {'new': 0, 'reopened': 0, 'resolved': 0, 'open': 0}
        if pk_fields:
            asd_mismatches = self._check_archive_source_diff(
                archive, domain, pk_fields, code_to_physical, field_name_map, stats)
            for m in asd_mismatches:
                m['check_type'] = ConsistencyIssue.CheckType.ARCHIVE_SOURCE_DIFF
                if not _is_rule_disabled(m['check_type'], m.get('field', ''), ''):
                    all_mismatches.append(m)
            asd_stats['open'] = len([m for m in asd_mismatches
                                     if not _is_rule_disabled(
                                         ConsistencyIssue.CheckType.ARCHIVE_SOURCE_DIFF,
                                         m.get('field', ''), '')])
        stats['by_type']['archive_source_diff'] = asd_stats

        # ===== 检查类型 3: orphan_source_record =====
        osr_stats = {'new': 0, 'reopened': 0, 'resolved': 0, 'open': 0}
        if pk_fields:
            osr_mismatches = self._check_orphan_source_records(
                archive, domain, pk_fields, code_to_physical, stats)
            for m in osr_mismatches:
                m['check_type'] = ConsistencyIssue.CheckType.ORPHAN_SOURCE_RECORD
                if not _is_rule_disabled(m['check_type'], '', ''):
                    all_mismatches.append(m)
            osr_stats['open'] = len([m for m in osr_mismatches
                                     if not _is_rule_disabled(
                                         ConsistencyIssue.CheckType.ORPHAN_SOURCE_RECORD, '', '')])
        stats['by_type']['orphan_source_record'] = osr_stats

        # ===== 检查类型 4: schema_drift =====
        sd_stats = {'new': 0, 'reopened': 0, 'resolved': 0, 'open': 0}
        sd_mismatches = self._check_schema_drift(archive, domain, field_name_map)
        for m in sd_mismatches:
            m['check_type'] = ConsistencyIssue.CheckType.SCHEMA_DRIFT
            if not _is_rule_disabled(m['check_type'], m.get('field', ''), ''):
                all_mismatches.append(m)
        sd_stats['open'] = len([m for m in sd_mismatches
                                if not _is_rule_disabled(
                                    ConsistencyIssue.CheckType.SCHEMA_DRIFT,
                                    m.get('field', ''), '')])
        stats['by_type']['schema_drift'] = sd_stats

        # ===== 汇总 upsert =====
        stats['mismatch_count'] = len(all_mismatches)
        stats['mismatch_records'] = len({m['record_key'] for m in all_mismatches})

        # 差异关联档案记录
        record_map = {}
        if pk_fields:
            for rec in ArchiveRecord.objects.filter(archive=archive).only('id', 'data'):
                k = '/'.join(str((rec.data or {}).get(pk, '')) for pk in pk_fields)
                if any(part for part in k.split('/')):
                    record_map.setdefault(k, rec.id)

        def _txt(v):
            return None if v is None else str(v)

        existing = {(i.record_key, i.field_code, i.member_source, i.check_type): i
                    for i in ConsistencyIssue.objects.filter(archive=archive)}
        seen, to_create, to_update = set(), [], []
        for m in all_mismatches:
            ct = m.get('check_type', ConsistencyIssue.CheckType.COMPOSITE_MEMBER)
            key = (m['record_key'][:200], m.get('field', ''), m.get('member_source', ''), ct)
            if key in seen:
                continue
            seen.add(key)
            issue = existing.get(key)
            if issue is None:
                to_create.append(ConsistencyIssue(
                    archive=archive, record_id=record_map.get(m['record_key']),
                    record_key=m['record_key'][:200], field_code=m.get('field', ''),
                    field_name=m.get('name', '') or field_name_map.get(m.get('field', ''), ''),
                    check_type=ct,
                    check_rule_key=m.get('check_rule_key', ''),
                    primary_source=m.get('primary_source', ''),
                    primary_value=_txt(m.get('primary_value')),
                    member_source=m.get('member_source', ''),
                    member_value=_txt(m.get('member_value')),
                    detail=m.get('detail'),
                    last_checked_at=now,
                ))
            else:
                issue.primary_value = _txt(m.get('primary_value'))
                issue.member_value = _txt(m.get('member_value'))
                issue.detail = m.get('detail', issue.detail)
                issue.record_id = issue.record_id or record_map.get(m['record_key'])
                issue.last_checked_at = now
                if issue.status == ConsistencyIssue.Status.RESOLVED:
                    issue.status = ConsistencyIssue.Status.OPEN
                    stats['reopened_issues'] += 1
                to_update.append(issue)
        ConsistencyIssue.objects.bulk_create(to_create)
        if to_update:
            ConsistencyIssue.objects.bulk_update(
                to_update, ['primary_value', 'member_value', 'record', 'last_checked_at',
                            'status', 'detail'])

        # 历史快照
        from .models import ConsistencyIssueHistory
        history_records = []
        if to_create:
            new_issues = ConsistencyIssue.objects.filter(
                archive=archive, last_checked_at=now
            ).exclude(id__in=[i.id for i in to_update])
            for issue in new_issues:
                history_records.append(ConsistencyIssueHistory(
                    issue=issue, checked_at=now,
                    primary_value=issue.primary_value, member_value=issue.member_value))
        for issue in to_update:
            history_records.append(ConsistencyIssueHistory(
                issue=issue, checked_at=now,
                primary_value=issue.primary_value, member_value=issue.member_value))
        if history_records:
            ConsistencyIssueHistory.objects.bulk_create(history_records)

        # 已消失的差异自动关闭
        if not stats['errors']:
            gone = [i for k, i in existing.items()
                    if k not in seen and i.status != ConsistencyIssue.Status.RESOLVED]
            for i in gone:
                i.status = ConsistencyIssue.Status.RESOLVED
                i.last_checked_at = now
            if gone:
                ConsistencyIssue.objects.bulk_update(gone, ['status', 'last_checked_at'])
            stats['resolved_issues'] = len(gone)

        stats['new_issues'] = len(to_create)
        stats['open_total'] = ConsistencyIssue.objects.filter(
            archive=archive, status=ConsistencyIssue.Status.OPEN).count()
        return Response(stats)

    @action(detail=True, methods=['post'], url_path='rollback-detail')
    def rollback_detail(self, request, pk=None):
        """重新同步明细子表：从源库拉全量数据覆盖，作为明细行回滚手段。
        不改变主表记录版本（代表行随明细同步自动更新）。

        POST /archives/{id}/rollback-detail/  body: {detail_fm_id: int, operated_by: str}
        """
        from apps.modeling.models import FieldMapping, Field as MField, Table

        archive = self.get_object()
        domain = archive.domain
        if not domain:
            return Response({'error': '档案未关联域'}, status=status.HTTP_400_BAD_REQUEST)

        detail_fm_id = request.data.get('detail_fm_id')
        operated_by = request.data.get('operated_by', 'system')
        if not detail_fm_id:
            return Response({'error': '必须提供 detail_fm_id'}, status=status.HTTP_400_BAD_REQUEST)

        detail_fm = get_object_or_404(
            FieldMapping, pk=detail_fm_id, relation_type=FieldMapping.RelationType.DETAIL)
        table = detail_fm.source_table
        if not table or not table.data_source:
            return Response({'error': '明细子表未配置数据源'}, status=status.HTTP_400_BAD_REQUEST)

        # 构造同步上下文（复用 _sync_data_from_sources 的公共逻辑）
        schema_type_map = {i['code']: i['type'] for i in (archive.schema or []) if i.get('code')}
        code_to_physical = self._build_code_to_physical(domain, schema_type_map)
        pk_fields = []
        primary_table = domain.get_primary_table()
        if primary_table:
            pk_fields = list(MField.objects.filter(
                table=primary_table, is_primary_key=True, status=MField.Status.ACTIVE
            ).values_list('code', flat=True))
        match_channels = {}
        for code, mappings in code_to_physical.items():
            seen = set()
            for tbl_id, phys in mappings:
                if tbl_id not in seen:
                    match_channels.setdefault(code, []).append((tbl_id, phys))
                    seen.add(tbl_id)

        stats = {'records_created': 0, 'records_updated': 0, 'records_deactivated': 0,
                 'records_reactivated': 0, 'details_created': 0, 'details_updated': 0,
                 'details_deactivated': 0, 'tables_synced': 0, 'errors': [],
                 'warnings': []}
        matched_ids = set()
        change_entries = []
        created_in_this_batch = set()
        sync_exclude_codes = set()

        try:
            rows = self._query_external_table(table)
            if rows is None:
                return Response({'error': f'明细子表 {table.name} 查询失败'}, status=500)
            self._sync_detail_rows(
                archive, table, rows, detail_fm, code_to_physical, pk_fields,
                match_channels, operated_by, stats, matched_ids, change_entries,
                created_in_this_batch, sync_exclude_codes,
            )
        except Exception as e:
            stats['errors'].append(str(e))
            return Response({'error': str(e)}, status=500)

        # 有变更 → 建批次
        if change_entries:
            batch = ArchiveChangeBatch.objects.create(
                archive=archive,
                change_source=ArchiveChangeBatch.ChangeSource.SYNC,
                operator=operated_by,
                stats={k: stats[k] for k in ('records_created', 'records_updated',
                                             'records_deactivated', 'records_reactivated',
                                             'details_created', 'details_updated',
                                             'details_deactivated') if k in stats},
            )
            from .serializers import _composite_label_codes, _build_record_label
            label_codes = _composite_label_codes(domain)
            rec_ids = [e['record_id'] for e in change_entries if e.get('record_id')]
            data_map = {}
            for i in range(0, len(rec_ids), 500):
                for r in ArchiveRecord.objects.filter(id__in=rec_ids[i:i + 500]).only('id', 'data'):
                    data_map[r.id] = r.data or {}
            ArchiveChangeDetail.objects.bulk_create([
                ArchiveChangeDetail(
                    batch=batch, archive=archive,
                    record_id=e.get('record_id'),
                    record_key=e.get('record_key', '')[:200],
                    record_label=_build_record_label(label_codes, data_map.get(e.get('record_id'))),
                    change_type=e['change_type'],
                    field_changes=e.get('field_changes', []),
                    version_before=e.get('version_before'),
                    version_after=e.get('version_after'),
                    detail_group_id=e.get('detail_group'),
                    detail_row_key=e.get('detail_row_key', ''),
                ) for e in change_entries
            ])
            stats['change_batch_id'] = batch.id

        ArchiveOperationLog.objects.create(
            archive=archive,
            operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.SYNC,
            change_summary={'action': f'明细子表重新同步（{table.name}）', 'detail_stats': stats},
        )
        return Response(stats)

    @action(detail=True, methods=['get'], url_path='permission-overview')
    def permission_overview(self, request, pk=None):
        """权限全景（仅管理员，只读审计视图）：一次聚合本档案的
        机器权限（API/暴露字段/授权密钥/调用统计）与人用权限（角色/字段授权/用户）。"""
        from apps.auth.views import IsMdmAdmin
        if not IsMdmAdmin().has_permission(request, self):
            return Response({'detail': '仅管理员可查看权限全景'},
                            status=status.HTTP_403_FORBIDDEN)
        archive = self.get_object()
        field_names = {i.get('code'): i.get('name') for i in (archive.schema or [])}

        # ── 机器权限：API + 密钥授权 + 调用统计（日志保留 90 天，按密钥聚合）──
        apis = []
        for api in ArchiveApi.objects.filter(archive=archive).order_by('id').prefetch_related('key_grants__api_key'):
            grants = [{
                'key_name': g.api_key.name,
                'key_status': g.api_key.status,
                'allowed_operations': g.allowed_operations or [],
            } for g in api.key_grants.all()]
            stats_by_key = {}
            total = 0
            last_at = None
            for log in ApiCallLog.objects.filter(api_id=api.id).iterator():
                total += 1
                if last_at is None or log.created_at > last_at:
                    last_at = log.created_at
                entry = stats_by_key.setdefault(
                    log.key_name or '(密钥已删除)',
                    {'count': 0, 'last_at': None, 'ips': set()})
                entry['count'] += 1
                if entry['last_at'] is None or log.created_at > entry['last_at']:
                    entry['last_at'] = log.created_at
                if log.client_ip and len(entry['ips']) < 5:
                    entry['ips'].add(log.client_ip)
            apis.append({
                'id': api.id,
                'name': api.name,
                'slug': api.slug,
                'status': api.status,
                'allowed_operations': api.allowed_operations or [],
                'exposed_fields': api.exposed_fields or [],
                'grants': grants,
                'call_stats': {
                    'total': total,
                    'last_at': last_at,
                    'by_key': [
                        {'key_name': k, 'count': v['count'],
                         'last_at': v['last_at'], 'ips': sorted(v['ips'])}
                        for k, v in sorted(stats_by_key.items(),
                                           key=lambda x: -x[1]['count'])
                    ],
                },
            })

        # ── 人用权限：角色×本域字段授权 + 挂靠用户 ──
        from apps.auth.models import RoleFieldPermission
        roles = []
        perms = RoleFieldPermission.objects.filter(
            domain_id=archive.domain_id).select_related('role').order_by('role_id')
        for perm in perms:
            role = perm.role
            users = [{
                'username': up.user.username,
                'display_name': up.display_name or up.user.username,
                'is_active': up.user.is_active,
            } for up in role.user_profiles.select_related('user').all()]
            roles.append({
                'role_id': role.id,
                'role_name': role.name,
                'is_builtin': role.is_builtin,
                'visible_codes': perm.visible_codes or [],
                'editable_codes': perm.editable_codes or [],
                'users': users,
            })

        return Response({
            'archive': {'id': archive.id, 'name': archive.name,
                        'domain_name': archive.domain.name},
            'field_names': field_names,
            'apis': apis,
            'roles': roles,
        })

    @action(detail=True, methods=['get'], url_path='field-distinct-values')
    def field_distinct_values(self, request, pk=None):
        """档案各字段去重值统计：从档案记录实时聚合每个 schema 字段的去重值及计数。"""
        archive = self.get_object()
        schema = archive.schema or []
        if not schema:
            return Response({'fields': [], 'total_records': 0})

        total_records = ArchiveRecord.objects.filter(
            archive=archive, status=ArchiveRecord.Status.ACTIVE
        ).count()

        # 按 schema code 收集去重值
        result = []
        for item in schema:
            code = item.get('code', '')
            if not code:
                continue
            # 从所有活跃记录的 data JSON 中提取该字段的值
            value_counts = {}
            for rec in ArchiveRecord.objects.filter(
                archive=archive, status=ArchiveRecord.Status.ACTIVE
            ).only('data').iterator():
                val = (rec.data or {}).get(code)
                if val is None or val == '':
                    continue
                # 统一转为字符串作为 key（JSON 值可能是 int/float/bool/str）
                key = str(val)
                value_counts[key] = value_counts.get(key, 0) + 1

            # 按计数降序排列，取前 200 个
            sorted_values = sorted(value_counts.items(), key=lambda x: -x[1])[:200]
            result.append({
                'code': code,
                'name': item.get('name', code),
                'group': item.get('group', ''),
                'type': item.get('type', 'string'),
                'distinct_count': len(value_counts),
                'values': [{'value': v, 'count': c} for v, c in sorted_values],
            })

        return Response({'fields': result, 'total_records': total_records})

    def _check_archive_source_diff(self, archive, domain, pk_fields, code_to_physical, field_name_map, stats):
        """检查类型 2: 档案侧人工覆盖与源侧数据差异。
        比对档案记录 manual_data 中的人工修改值与源表当前值。
        """
        from apps.modeling.models import Table, Field

        mismatches = []
        schema = archive.schema or []
        # 找出档案维护(ownership=archive)或有修正保护的字段
        manual_owned = {i['code'] for i in schema
                        if i.get('ownership') == 'archive' and i.get('source') != 'computed'}

        if not manual_owned:
            return mismatches

        # 采集源表数据（按主键索引）
        source_by_pk = {}
        for table in Table.objects.filter(domain=domain, status=Table.Status.ACTIVE):
            try:
                rows = self._query_local_table(table) if not table.data_source \
                    else self._query_external_table(table)
                if rows is None:
                    continue
                # 构建物理列→schema code 映射
                phys_to_schema = {}
                for sc, mappings in code_to_physical.items():
                    for tbl_id, pc in mappings:
                        if tbl_id == table.id:
                            phys_to_schema[pc] = sc
                for row in rows:
                    record_data = {}
                    for col_name, value in row.items():
                        sc = phys_to_schema.get(col_name, col_name)
                        record_data[sc] = value
                    key = tuple(str(record_data.get(pk, '')) for pk in pk_fields)
                    if any(k for k in key):
                        source_by_pk[key] = record_data
            except Exception as e:
                stats['errors'].append(f'archive_source_diff/{table.name}: {str(e)}')

        # 比对档案记录
        for rec in ArchiveRecord.objects.filter(archive=archive, status='active'):
            rec_key = '/'.join(str((rec.data or {}).get(pk, '')) for pk in pk_fields)
            if not any(part for part in rec_key.split('/')):
                continue
            overrides = rec.overrides or {}
            source_row = source_by_pk.get(tuple(rec_key.split('/')))
            if not source_row:
                continue
            for code in manual_owned:
                if code not in overrides:
                    continue  # 没有人工覆盖的不检查
                archive_val = (rec.manual_data or {}).get(code)
                source_val = source_row.get(code)
                if archive_val is not None and str(archive_val) != str(source_val) if source_val is not None else True:
                    mismatches.append({
                        'record_key': rec_key,
                        'field': code,
                        'name': field_name_map.get(code, code),
                        'primary_source': f'档案人工覆盖',
                        'primary_value': archive_val,
                        'member_source': f'源侧数据',
                        'member_value': source_val,
                        'check_rule_key': f'archive_source_diff:{code}',
                        'detail': {'archive_record_id': rec.id, 'override_info': overrides.get(code)},
                    })
        return mismatches

    def _check_orphan_source_records(self, archive, domain, pk_fields, code_to_physical, stats):
        """检查类型 3: 源侧模型中存在没有关联上主表主键的数据。
        即源表中有数据但其主键值在档案中找不到对应记录。
        """
        from apps.modeling.models import Table

        mismatches = []
        # 采集档案中已有的主键值集合
        archive_pk_set = set()
        for rec in ArchiveRecord.objects.filter(archive=archive).only('data'):
            k = tuple(str((rec.data or {}).get(pk, '')) for pk in pk_fields)
            if any(part for part in k):
                archive_pk_set.add(k)

        # 遍历源表，找不在档案中的记录
        primary_table = domain.get_primary_table()
        tables_to_check = Table.objects.filter(domain=domain, status=Table.Status.ACTIVE)
        if primary_table:
            tables_to_check = tables_to_check.filter(id=primary_table.id)

        for table in tables_to_check:
            try:
                rows = self._query_local_table(table) if not table.data_source \
                    else self._query_external_table(table)
                if rows is None:
                    continue
                # 构建物理列→schema code 映射
                phys_to_schema = {}
                for sc, mappings in code_to_physical.items():
                    for tbl_id, pc in mappings:
                        if tbl_id == table.id:
                            phys_to_schema[pc] = sc
                for row in rows:
                    record_data = {}
                    for col_name, value in row.items():
                        sc = phys_to_schema.get(col_name, col_name)
                        record_data[sc] = value
                    key = tuple(str(record_data.get(pk, '')) for pk in pk_fields)
                    if not any(k for k in key):
                        continue
                    if key not in archive_pk_set:
                        pk_display = '/'.join(key)
                        mismatches.append({
                            'record_key': pk_display,
                            'field': '',
                            'name': '',
                            'primary_source': '',
                            'primary_value': pk_display,
                            'member_source': f'{table.name}',
                            'member_value': pk_display,
                            'check_rule_key': f'orphan_source_record:{table.name}',
                            'detail': {'table': table.name, 'pk_values': dict(zip(pk_fields, key)),
                                       'sample_data': {k: v for k, v in list(record_data.items())[:5]}},
                        })
            except Exception as e:
                stats['errors'].append(f'orphan/{table.name}: {str(e)}')
        return mismatches

    def _check_schema_drift(self, archive, domain, field_name_map):
        """检查类型 4: 档案 schema 与当前建模结构不一致。
        检测 schema 中的字段在当前建模中是否还存在、类型是否匹配。
        """
        from apps.modeling.models import StandardField, Field, Table

        mismatches = []
        schema = archive.schema or []
        if not schema:
            return mismatches

        # 当前建模中的所有标准字段
        current_sf = {}
        for sf in StandardField.objects.filter(domain=domain, status='active', is_active=True):
            current_sf[sf.standard_code] = sf

        # 当前建模中的所有物理字段
        current_fields = {}
        for f in Field.objects.filter(table__domain=domain, status=Field.Status.ACTIVE):
            current_fields[f.code or f.name] = f

        for item in schema:
            code = item.get('code', '')
            if not code:
                continue
            source = item.get('source', '')
            if source == 'computed':
                continue  # 计算字段不参与漂移检查

            if code not in current_sf and code not in current_fields:
                # schema 中的字段在建模中已不存在
                mismatches.append({
                    'record_key': '',
                    'field': code,
                    'name': field_name_map.get(code, code),
                    'primary_source': '档案 schema',
                    'primary_value': f'{item.get("type", "?")} ({source})',
                    'member_source': '建模结构',
                    'member_value': '字段已不存在',
                    'check_rule_key': f'schema_drift:{code}:removed',
                    'detail': {'issue': 'field_removed', 'schema_item': item},
                })
            elif code in current_sf:
                sf = current_sf[code]
                # 检查类型是否变化
                schema_type = item.get('type', '')
                if schema_type and sf.field_type and schema_type != sf.field_type:
                    mismatches.append({
                        'record_key': '',
                        'field': code,
                        'name': field_name_map.get(code, code),
                        'primary_source': '档案 schema',
                        'primary_value': schema_type,
                        'member_source': '建模结构',
                        'member_value': sf.field_type,
                        'check_rule_key': f'schema_drift:{code}:type_change',
                        'detail': {'issue': 'type_changed', 'schema_type': schema_type,
                                   'modeling_type': sf.field_type},
                    })
        return mismatches

    def _preview_data_changes(self, archive, domain, new_schema):
        """数据变化试算（只读不写）：按与 _sync_data_from_sources 同口径拉源行，
        跨表累积后用 _merge_record_data 模拟合并，统计将新增/更新/停用的记录数及字段变化样本。"""
        from types import SimpleNamespace
        from apps.modeling.models import Table, Field

        schema_type_map = {i['code']: i['type'] for i in new_schema if i.get('code')}
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in new_schema}
        # v18 预检告警：档案维护字段（非 source）被源侧刷新波及的统计
        archive_owned_codes = {i['code'] for i in new_schema
                               if i.get('code') and (i.get('ownership') or 'archive') != 'source'}
        code_to_physical = self._build_code_to_physical(domain, schema_type_map)
        stats = {'tables_checked': 0, 'would_create': 0, 'would_update': 0,
                 'would_deactivate': 0, 'changes_sample': [], 'errors': [], 'warnings': [],
                 'archive_owned_impact': {'records': 0, 'fields': []}}

        # 主字段告警（不阻断）：组合字段未设主字段时仅告警，其余字段正常同步
        missing_pf = _validate_primary_fields(domain)
        if missing_pf:
            names = '、'.join(f"{m['name']}({m['code']})" for m in missing_pf)
            stats['warnings'].append(f'以下组合字段未设置主字段，其成员数据将全部写入（建议到属性配置页设置）：{names}')
            stats['primary_field_missing'] = missing_pf

        primary_table = domain.get_primary_table()
        pk_fields = []
        if primary_table:
            for f in Field.objects.filter(table=primary_table, is_primary_key=True, status=Field.Status.ACTIVE):
                sf = f.standard_field
                if sf:
                    pk_fields.append(sf.standard_code)
                else:
                    pk_fields.append(f.code)
        if not pk_fields:
            first_pk = Field.objects.filter(
                table__domain=domain, is_primary_key=True, status=Field.Status.ACTIVE
            ).first()
            if first_pk:
                sf = first_pk.standard_field
                pk_fields = [sf.standard_code if sf else first_pk.code]
        if not pk_fields:
            stats['errors'].append('未配置主键字段，无法试算')
            return stats

        all_tables = Table.objects.filter(domain=domain, status=Table.Status.ACTIVE)
        if primary_table:
            tables = [primary_table] + [t for t in all_tables if t.id != primary_table.id]
        else:
            tables = list(all_tables)

        # 主键索引（与正式同步同规则：同键 active 优先）
        existing_records = {}
        for rec in ArchiveRecord.objects.filter(archive=archive).order_by('id'):
            key = tuple(str(rec.data.get(pk, '')) for pk in pk_fields)
            if not any(k for k in key):
                continue
            prev = existing_records.get(key)
            if prev is not None and prev.status == ArchiveRecord.Status.ACTIVE:
                continue
            if prev is None or rec.status == ArchiveRecord.Status.ACTIVE:
                existing_records[key] = rec

        new_keys = set()
        updates_by_key = {}  # 跨表累积已有记录的源字段更新
        for table in tables:
            # ===== 跳过预组合/明细子表（与 _sync_data_from_sources 同口径）：
            # 此类表数据不入主记录（ArchiveRecord），走 _sync_detail_rows 分支 =====
            if table.data_source:
                from apps.modeling.models import DetailTableConfig, FieldMapping
                cfg = DetailTableConfig.objects.filter(domain=domain, table=table).first()
                if cfg:
                    continue  # 有子表注册，数据入 ArchiveRecordDetail
                old_fm = FieldMapping.objects.filter(
                    source_table=table,
                    relation_type=FieldMapping.RelationType.DETAIL,
                ).first()
                if old_fm:
                    continue  # 旧内嵌 detail 配置兼容
            try:
                rows = self._query_local_table(table) if not table.data_source else self._query_external_table(table)
            except Exception as e:
                stats['errors'].append(f'{table.name or table.code}: {e}')
                continue
            if rows is None:
                continue
            stats['tables_checked'] += 1
            physical_to_schema = {}
            for schema_code, mappings in code_to_physical.items():
                for tbl_id, phys_col in mappings:
                    if tbl_id == table.id:
                        physical_to_schema[phys_col] = schema_code
            for row in rows:
                record_data = {}
                for col_name, value in row.items():
                    sc = physical_to_schema.get(col_name)
                    # 同名兜底仅限主键列（记录匹配必需）：未映射给本表的其他同名列
                    # 不得写入，否则他表同名空列会偷渡清空已有值造成假变更（BUG-2026-0805-01）
                    if not sc and col_name in pk_fields and col_name in schema_type_map:
                        sc = col_name
                    if sc:
                        record_data[sc] = value
                if not record_data:
                    continue
                key = tuple(str(record_data.get(pk, '')) for pk in pk_fields)
                if not any(k for k in key):
                    continue
                if key in existing_records:
                    updates_by_key.setdefault(key, {}).update(record_data)
                else:
                    new_keys.add(key)

        stats['would_create'] = len(new_keys)
        matched_ids = set()
        impact_keys = set()      # v18：被波及的记录主键
        impact_fields = {}       # v18：档案维护字段 → 波及记录数
        for key, record_data in updates_by_key.items():
            existing = existing_records[key]
            matched_ids.add(existing.id)
            sim = SimpleNamespace(
                source_data={**(existing.source_data or {}), **record_data},
                manual_data=dict(existing.manual_data or {}),
                data=existing.data or {},
                lineage=dict(existing.lineage or {}),
            )
            merged, _ = _merge_record_data(sim, new_schema)
            old_data = existing.data or {}
            changed_codes = sorted(
                c for c in set(list(merged.keys()) + list(old_data.keys()))
                if old_data.get(c) != merged.get(c)
            )
            if changed_codes:
                stats['would_update'] += 1
                # v18：源侧刷新波及档案维护字段（无人工覆盖层时值取自源层）→ 预检告警素材
                owned_hit = [c for c in changed_codes if c in archive_owned_codes]
                if owned_hit:
                    impact_keys.add(key)
                    for c in owned_hit:
                        impact_fields[c] = impact_fields.get(c, 0) + 1
                if len(stats['changes_sample']) < 20:
                    stats['changes_sample'].append({
                        'record_key': '/'.join(k for k in key),
                        'changed_fields': [
                            {'field': c, 'name': field_name_map.get(c, c),
                             'old': old_data.get(c), 'new': merged.get(c)}
                            for c in changed_codes[:10]
                        ],
                    })

        if stats['tables_checked'] > 0 and not stats['errors']:
            stats['would_deactivate'] = ArchiveRecord.objects.filter(
                archive=archive, status=ArchiveRecord.Status.ACTIVE,
                sync_status__in=['synced', 'partial'],
            ).exclude(id__in=matched_ids).count()
        # v18 预检告警：档案维护字段被源侧刷新波及（提醒用户确认，不阻断刷新）
        stats['archive_owned_impact'] = {
            'records': len(impact_keys),
            'fields': [{'code': c, 'name': field_name_map.get(c, c), 'records': n}
                       for c, n in sorted(impact_fields.items(), key=lambda x: -x[1])],
        }
        return stats

    def _sync_data_from_sources(self, archive, domain, schema_type_map, operated_by):
        """从域的数据源表拉取数据，创建/更新档案记录。
        
        架构逻辑：
        1. 获取域的主表（is_primary=True）
        2. 获取主表的主键字段（is_primary_key=True）
        3. 先处理主表数据，创建记录
        4. 用主键字段匹配，合并其他表数据
        """
        from apps.modeling.models import Table, Field, StandardField, FieldMapping
        from apps.modeling.views import DataSourceViewSet, _json_safe
        from django.db import connections

        stats = {'records_created': 0, 'records_updated': 0, 'tables_synced': 0,
                 'records_deactivated': 0, 'records_reactivated': 0,
                 'details_created': 0, 'details_updated': 0, 'details_deactivated': 0,
                 'errors': [], 'warnings': []}
        # 本轮刷新中匹配到源行的记录 id（跨表共享，用于收尾停用清扫）
        matched_ids = set()
        # 本轮变更明细（源侧同步，收尾统一落变更日志批次）
        change_entries = []
        # 本轮同步中刚创建的记录 ID（后续表同步到这些记录时不计入「修改」）
        created_in_this_batch = set()

        # ===== 主字段告警（不阻断）：组合字段未设主字段时仅告警，其余字段正常同步 =====
        missing_pf = _validate_primary_fields(domain)
        if missing_pf:
            names = '、'.join(f"{m['name']}({m['code']})" for m in missing_pf)
            stats['warnings'].append(f'以下组合字段未设置主字段，其成员数据将全部写入（建议到属性配置页设置）：{names}')
            stats['primary_field_missing'] = missing_pf

        # 获取主表及其主键字段（用 schema code 而非 Field.code，因为 record_data 以 schema code 为 key）
        primary_table = domain.get_primary_table()
        pk_fields = []
        if primary_table:
            for f in Field.objects.filter(table=primary_table, is_primary_key=True, status=Field.Status.ACTIVE):
                sf = f.standard_field
                if sf:
                    pk_fields.append(sf.standard_code)
                else:
                    pk_fields.append(f.code)
        
        # 如果没有主表或主键字段，使用默认的第一个字段作为匹配键
        if not pk_fields:
            # 回退：使用所有表的第一个主键字段
            first_pk = Field.objects.filter(
                table__domain=domain, is_primary_key=True, status=Field.Status.ACTIVE
            ).first()
            if first_pk:
                sf = first_pk.standard_field
                pk_fields = [sf.standard_code if sf else first_pk.code]
        
        all_tables = Table.objects.filter(domain=domain, status=Table.Status.ACTIVE)
        
        # 主表优先处理
        if primary_table:
            tables = [primary_table] + [t for t in all_tables if t.id != primary_table.id]
        else:
            tables = list(all_tables)

        # 构建 schema code → 物理字段映射
        code_to_physical = self._build_code_to_physical(domain, schema_type_map)
        # 匹配通道：组合字段非主成员列（外键）仅用于构建记录 key，不写入
        match_channels = self._build_match_channels(domain, pk_fields)

        # 一致性检查准备：非主字段成员只检查不写入
        code_checks = self._build_code_checks(domain)
        cc_primary_values = {}
        cc_member_values = {}
        # 构建同步排除集合：组合字段的非主字段成员不写入档案（只用于一致性检查）
        sync_exclude_codes = self._build_sync_exclude_codes(code_checks, code_to_physical)

        # 各表确定性排序列（主键物理列，1:n 取首条折叠用）
        pk_col_by_table = {}
        for t in tables:
            f = Field.objects.filter(table=t, is_primary_key=True, status=Field.Status.ACTIVE).first()
            if f:
                pk_col_by_table[t.id] = f.physical_name or f.code or f.name

        def _has_direct_pk_mapping(t):
            """该表是否有物理列可直接映射到档案主键（直接匹配 vs 经 FieldMapping 中转）。"""
            for pk in pk_fields:
                for tbl_id, phys in list(code_to_physical.get(pk, [])) + list(match_channels.get(pk, [])):
                    if tbl_id == t.id:
                        return True
            return False

        is_primary_id = primary_table.id if primary_table else None

        # 预组合过滤预扫（2026-08-14）：inner detail 挂载条件交集 → 主记录行级过滤
        # （主表优先全量入档后无法收敛——主表 upsert 必须发生在过滤之后）
        precombine_filters = self._build_precombine_filters(
            domain, tables, pk_fields, code_to_physical, match_channels, stats)

        # 处理每个表
        for table in tables:
            is_primary_table = (table.id == is_primary_id)
            # 仅非主表排序（1:n 折叠确定性）；主表保持原查询语义，避免 ORDER BY 改变 TOP 截断集合
            order_by = pk_col_by_table.get(table.id) if not is_primary_table else None
            if not table.data_source:
                try:
                    rows = self._query_local_table(table)
                    if rows is not None:
                        self._upsert_records_from_rows(
                            archive, table, rows, code_to_physical, schema_type_map, 
                            pk_fields, operated_by, stats, matched_ids, change_entries, created_in_this_batch,
                            sync_exclude_codes, match_channels, is_primary_table,
                            row_filter=precombine_filters.get(table.id),
                        )
                        self._collect_check_values(table, rows, code_checks, pk_fields,
                                                   code_to_physical, cc_primary_values, cc_member_values)
                        stats['tables_synced'] += 1
                except Exception as e:
                    stats['errors'].append(f'本地表 {table.name}: {str(e)}')
                continue

            # ===== 子表关系分支（2026-08-08/08-11）：该表配置 relation_type=detail 的 FieldMapping 时，
            # 整表作为明细致子表同步（保留全部行），不进入直连/中转路径。
            # 2026-08-11 多挂载改造：先查 DetailTableConfig（先注册后挂载），
            # 有则循环多挂载；无注册则回退查询旧内嵌配置兼容。
            # =====
            detail_fms = []
            if table.data_source:
                from apps.modeling.models import DetailTableConfig
                cfg = DetailTableConfig.objects.filter(domain=archive.domain, table=table).first()
                if cfg:
                    # 子表已注册：查出挂载的全部 detail 映射（一子表多挂载）
                    detail_fms = list(FieldMapping.objects.filter(
                        detail_config=cfg,
                        source_field__status='active',
                    ).select_related('row_key_field', 'display_sort_field', 'detail_config'))
                else:
                    # 存量兼容：查旧内嵌 detail 配置（deprecated）
                    old_fm = FieldMapping.objects.filter(
                        source_table=table,
                        relation_type=FieldMapping.RelationType.DETAIL,
                        source_field__status='active',
                    ).select_related('row_key_field', 'display_sort_field').first()
                    if old_fm:
                        stats['warnings'].append(
                            f'明细子表 {table.name}：使用旧内嵌配置（未注册子表配置），'
                            f'建议在关系管理注册子表后再挂载')
                        detail_fms = [old_fm]
            if detail_fms:
                # TODO（2026-08-11 性能优化）：多挂载时 rows 只拉一次，多路复用
                # conditions 从第一个挂载的 detail_config（或 fm 自身 deprecated）取
                first_fm = detail_fms[0]
                cfg = first_fm.detail_config
                conds = None
                if cfg and cfg.conditions:
                    conds = cfg.conditions
                elif first_fm.conditions:
                    conds = first_fm.conditions
                # 2026-08-14：条件按 field_source 拆分——header 条件应用到预组合头表查询
                # （header 字段不在明细表白名单，原实现 ValueError 整表跳过=筛选条件静默失效）
                header_conds, detail_conds = self._split_conditions(conds)
                rows = self._query_external_table(table, order_by=order_by, conditions=detail_conds)
                if rows is not None:
                    # 预组合平铺（2026-08-11 第三轮）：注册配了头表时，头表全量 JOIN 进明细行
                    # （头表字段以 __hdr__ 前缀并入，一行=一条明细+头字段重复）
                    if cfg and cfg.header_table_id and cfg.header_link_field_id and cfg.detail_link_field_id:
                        rows = self._join_header_rows(table, cfg, rows, first_fm.join_type,
                                                      conditions=header_conds)
                    for detail_fm in detail_fms:
                        try:
                            self._sync_detail_rows(
                                archive, table, rows, detail_fm, code_to_physical,
                                pk_fields, match_channels, operated_by, stats, matched_ids,
                                change_entries, created_in_this_batch, sync_exclude_codes,
                            )
                        except Exception as e:
                            stats['errors'].append(f'数据源表 {table.name}（明细子表）: {str(e)}')
                    self._collect_check_values(table, rows, code_checks, pk_fields,
                                               code_to_physical, cc_primary_values, cc_member_values)
                    stats['tables_synced'] += 1
                continue

            try:
                if _has_direct_pk_mapping(table):
                    rows = self._query_external_table(table, order_by=order_by)
                    if rows is not None:
                        self._upsert_records_from_rows(
                            archive, table, rows, code_to_physical, schema_type_map,
                            pk_fields, operated_by, stats, matched_ids, change_entries, created_in_this_batch,
                            sync_exclude_codes, match_channels, is_primary_table,
                            row_filter=precombine_filters.get(table.id),
                        )
                        self._collect_check_values(table, rows, code_checks, pk_fields,
                                                   code_to_physical, cc_primary_values, cc_member_values)
                        stats['tables_synced'] += 1
                else:
                    # 维度表（无主键直映）：经 FieldMapping 中转匹配主记录（如价目表 NAME 经明细 FID 关联物料）
                    rows = self._query_external_table(table, order_by=order_by)
                    if rows is not None:
                        self._upsert_dimension_via_mapping(
                            archive, table, rows, code_to_physical, schema_type_map,
                            pk_fields, match_channels, operated_by, stats, matched_ids,
                            change_entries, created_in_this_batch
                        )
                        stats['tables_synced'] += 1
            except Exception as e:
                stats['errors'].append(f'数据源表 {table.name}: {str(e)}')

        # ===== 一致性检查：非主字段成员值 vs 主字段值（告警不阻断） =====
        if code_checks:
            field_name_map = {i.get('code'): i.get('name') or i.get('code')
                              for i in (archive.schema or [])}
            stats['consistency_check'] = self._run_consistency_check(
                code_checks, cc_primary_values, cc_member_values, field_name_map)

        # ===== 停用清扫：源侧已删除的记录标记停用（只标不删） =====
        # 安全闸门：任一表同步出错或无主键时跳过，防止源库瞬时故障引发误停用
        if stats['tables_synced'] > 0 and not stats['errors'] and pk_fields:
            # matched_ids 可达 20 万+：先拉候选 id 集与 matched 求差（内存），再按 500/批
            # id__in 分批处理。禁止用多次 exclude(id__in=...)——AND 嵌套 NOT IN 仍超 SQLite
            # 999 变量上限（实测 too many SQL variables，见 BUG-2026-0808-02）
            candidate_ids = set(ArchiveRecord.objects.filter(
                archive=archive, status=ArchiveRecord.Status.ACTIVE,
                sync_status__in=['synced', 'partial'],
            ).values_list('id', flat=True))
            stale_ids = sorted(candidate_ids - matched_ids)
            # 先抓取待停用记录身份（变更日志用），再批量更新
            for i in range(0, len(stale_ids), 500):
                chunk = stale_ids[i:i + 500]
                for rec in ArchiveRecord.objects.filter(id__in=chunk).only('id', 'data', 'version'):
                    change_entries.append({
                        'record_id': rec.id,
                        'record_key': '/'.join(str(rec.data.get(pk, '')) for pk in pk_fields),
                        'change_type': ArchiveChangeDetail.ChangeType.DEACTIVATED,
                        'field_changes': [],
                        # 停用清扫不动数据层，无版本变动（v18 版本映射）
                        'version_before': rec.version,
                        'version_after': rec.version,
                    })
                ArchiveRecord.objects.filter(id__in=chunk).update(
                    status=ArchiveRecord.Status.DELETED, sync_status='stale',
                    updated_by=operated_by, updated_at=timezone.now(),
                )
            stats['records_deactivated'] = len(stale_ids)

        # ===== 变更日志：本轮有变更才建批次（零变更不产生噪声） =====
        if change_entries:
            batch = ArchiveChangeBatch.objects.create(
                archive=archive,
                change_source=ArchiveChangeBatch.ChangeSource.SYNC,
                operator=operated_by,
                stats={k: stats[k] for k in ('records_created', 'records_updated',
                                             'records_deactivated', 'records_reactivated')},
            )
            # 记录信息快照：组合字段值拼接，让用户一眼识别变更的是哪条数据
            from .serializers import _composite_label_codes, _build_record_label
            label_codes = _composite_label_codes(domain)
            rec_ids = [e['record_id'] for e in change_entries if e.get('record_id')]
            # id 列表可达 20 万+，分批查询避免 SQLite 变量上限（与停用清扫同口径）
            data_map = {}
            for i in range(0, len(rec_ids), 500):
                for r in ArchiveRecord.objects.filter(id__in=rec_ids[i:i + 500]).only('id', 'data'):
                    data_map[r.id] = r.data or {}
            ArchiveChangeDetail.objects.bulk_create([
                ArchiveChangeDetail(
                    batch=batch, archive=archive,
                    record_id=e.get('record_id'),
                    record_key=e.get('record_key', '')[:200],
                    record_label=_build_record_label(label_codes, data_map.get(e.get('record_id'))),
                    change_type=e['change_type'],
                    field_changes=e.get('field_changes', []),
                    version_before=e.get('version_before'),
                    version_after=e.get('version_after'),
                    detail_group_id=e.get('detail_group'),
                    detail_row_key=e.get('detail_row_key', ''),
                ) for e in change_entries
            ])
            stats['change_batch_id'] = batch.id

        return stats

    def _build_code_to_physical(self, domain, schema_type_map):
        """构建 schema code → [(table_id, 物理列名)] 写入映射（正式同步与刷新预检共用）。

        组合字段（StandardField）已设主字段时仅映射主字段成员（数据源头唯一）；
        其余成员不再写底层，只参与一致性检查（_build_code_checks）。
        """
        from apps.modeling.models import Table, Field, StandardField

        code_to_physical = {}
        primary_locked = set()  # 已按主字段锁定映射的 sf_id，后续兜底循环不再追加成员
        standard_fields = StandardField.objects.filter(domain=domain).prefetch_related('members__table')
        for sf in standard_fields:
            members = [m for m in sf.members.all()
                       if m.table and m.table.status == Table.Status.ACTIVE]
            if not members:
                continue
            pf = next((m for m in members if m.id == sf.primary_field_id and m.status == 'active'), None)
            if pf is not None:
                # 主字段作为唯一数据源头（用 physical_name 保留原始列名，改名不影响同步）
                code_to_physical[sf.standard_code] = [(pf.table_id, pf.physical_name or pf.code or pf.name)]
                primary_locked.add(sf.id)
            else:
                code_to_physical[sf.standard_code] = [(m.table_id, m.physical_name or m.code or m.name) for m in members]

        all_fields = Field.objects.filter(
            table__domain=domain, status=Field.Status.ACTIVE
        ).select_related('table')
        # solo 字段：按 Field.code 匹配 schema code，用 physical_name 作为物理列名
        for f in all_fields:
            if f.code in schema_type_map and f.code not in code_to_physical:
                code_to_physical[f.code] = [(f.table_id, f.physical_name or f.code or f.name)]
        for f in all_fields:
            if f.standard_field_id and f.standard_field_id not in primary_locked:
                sf_code = None
                for sf in standard_fields:
                    if sf.id == f.standard_field_id:
                        sf_code = sf.standard_code
                        break
                if sf_code and sf_code in schema_type_map:
                    existing = code_to_physical.get(sf_code, [])
                    entry = (f.table_id, f.physical_name or f.code or f.name)
                    if entry not in existing:
                        existing.append(entry)
                        code_to_physical[sf_code] = existing

        for code in schema_type_map:
            if code not in code_to_physical:
                code_to_physical[code] = []
        return code_to_physical

    def _build_match_channels(self, domain, pk_fields):
        """构建主键匹配通道：pk_fields 对应组合字段的全部成员列（含非主成员）。

        返回 {schema_code: [(table_id, physical), ...]}。
        维度表（如价目明细）的外键列仅用于构建记录匹配 key，不写入档案（防多源覆盖）。
        """
        from apps.modeling.models import StandardField, Field, Table

        channels = {}
        for pk_code in pk_fields:
            entries = []
            sf = StandardField.objects.filter(domain=domain, standard_code=pk_code).first()
            if sf:
                for m in sf.members.all():
                    if m.table and m.table.status == Table.Status.ACTIVE:
                        entries.append((m.table_id, m.physical_name or m.code or m.name))
            # solo 字段兜底（无组合字段时按 code 直配）
            for f in Field.objects.filter(
                code=pk_code, status=Field.Status.ACTIVE, table__domain=domain
            ):
                entry = (f.table_id, f.physical_name or f.code or f.name)
                if entry not in entries:
                    entries.append(entry)
            if entries:
                channels[pk_code] = entries
        return channels

    def _build_code_checks(self, domain):
        """构建一致性检查映射：已设主字段的组合字段，非主字段成员的取值与主字段比对。

        返回 {schema_code: {'primary': (table_id, 物理列, 表名),
                          'others': [(table_id, 物理列, 表名), ...]}}
        """
        from apps.modeling.models import Table, StandardField

        checks = {}
        qs = StandardField.objects.filter(
            domain=domain, status='active', is_active=True
        ).prefetch_related('members__table')
        for sf in qs:
            members = [m for m in sf.members.all()
                       if m.status == 'active' and m.table and m.table.status == Table.Status.ACTIVE]
            pf = next((m for m in members if m.id == sf.primary_field_id), None)
            if pf is None:
                continue
            others = [(m.table_id, m.code or m.name, m.table.name) for m in members if m.id != pf.id]
            if others:
                checks[sf.standard_code] = {
                    'primary': (pf.table_id, pf.code or pf.name, pf.table.name),
                    'others': others,
                }
        return checks

    def _build_sync_exclude_codes(self, code_checks, code_to_physical):
        """构建同步排除集合：组合字段的非主字段成员不写入档案。

        返回 {(table_id, physical_column), ...} 集合，
        _upsert_records_from_rows 遇到这些字段时跳过不写入。
        """
        exclude = set()
        for schema_code, check_info in code_checks.items():
            # 非主字段成员的 (table_id, physical_column) 加入排除集合
            for table_id, phys_col, _ in check_info.get('others', []):
                exclude.add((table_id, phys_col))
        return exclude

    def _collect_check_values(self, table, rows, code_checks, pk_fields, code_to_physical,
                              primary_values, member_values):
        """从单表拉取结果中采集一致性检查所需的主字段值/成员值（按主键分组）。"""
        if not code_checks:
            return
        # 本表 物理列 → schema code（主键提取用，与 upsert 同口径）
        physical_to_schema = {}
        for schema_code, mappings in code_to_physical.items():
            for tbl_id, phys_col in mappings:
                if tbl_id == table.id:
                    physical_to_schema[phys_col] = schema_code
        for row in rows:
            record_data = {}
            for col_name, value in row.items():
                sc = physical_to_schema.get(col_name, col_name)
                record_data[sc] = value
            key = tuple(str(record_data.get(pk, '')) for pk in pk_fields)
            if not any(k for k in key):
                continue
            for code, info in code_checks.items():
                p_tbl, p_col, _ = info['primary']
                if p_tbl == table.id and p_col in row:
                    primary_values.setdefault(key, {})[code] = row[p_col]
                for o_tbl, o_col, o_label in info['others']:
                    if o_tbl == table.id and o_col in row:
                        member_values.setdefault(key, {}).setdefault(code, []).append((o_label, o_col, row[o_col]))

    def _run_consistency_check(self, code_checks, primary_values, member_values, field_name_map):
        """比对非主字段成员值与主字段值，生成告警报告（不阻断，数据仍以主字段为准）。"""
        def _norm(v):
            return '' if v is None else str(v)

        mismatch_records = set()
        mismatch_count = 0
        samples = []
        for key, codes in member_values.items():
            for code, entries in codes.items():
                p_val = (primary_values.get(key) or {}).get(code)
                for (label, col, val) in entries:
                    if _norm(val) != _norm(p_val):
                        mismatch_count += 1
                        mismatch_records.add(key)
                        if len(samples) < 20:
                            p_info = code_checks.get(code, {}).get('primary')
                            samples.append({
                                'record_key': '/'.join(key),
                                'field': code,
                                'name': field_name_map.get(code, code),
                                'primary_source': f'{p_info[2]}.{p_info[1]}' if p_info else '',
                                'primary_value': p_val,
                                'member_source': f'{label}.{col}',
                                'member_value': val,
                            })
        return {
            'checked_fields': len(code_checks),
            'mismatch_count': mismatch_count,
            'mismatch_records': len(mismatch_records),
            'samples': samples,
            'checked_at': timezone.now().isoformat(),
        }

    def _collect_full_mismatches(self, code_checks, primary_values, member_values, field_name_map):
        """全量比对（一致性检查页专用，不截断样本），返回差异明细列表。"""
        def _norm(v):
            return '' if v is None else str(v)

        out = []
        for key, codes in member_values.items():
            for code, entries in codes.items():
                p_val = (primary_values.get(key) or {}).get(code)
                p_info = code_checks.get(code, {}).get('primary')
                for (label, col, val) in entries:
                    if _norm(val) != _norm(p_val):
                        out.append({
                            'record_key': '/'.join(key),
                            'field': code,
                            'name': field_name_map.get(code, code),
                            'primary_source': f'{p_info[2]}.{p_info[1]}' if p_info else '',
                            'primary_value': p_val,
                            'member_source': f'{label}.{col}',
                            'member_value': val,
                        })
        return out

    def _query_local_table(self, table):
        """查询本地表数据"""
        from django.db import connection
        try:
            with connection.cursor() as cursor:
                cursor.execute(f'SELECT * FROM "{table.code}"')
                columns = [desc[0] for desc in cursor.description]
                rows = []
                for row in cursor.fetchall():
                    rows.append({col: _json_safe(val) for col, val in zip(columns, row)})
                return rows
        except Exception:
            return None

    def _query_external_table(self, table, order_by=None, conditions=None, count_only=False):
        """查询外部数据源表数据（全量，无行数截断）。order_by 为物理列名时附加确定性排序（取首条折叠用）。

        conditions（2026-08-08 新增）：结构化 ON/WHERE 筛选条件（AND 组合），
        每项 {"field": 物理列名或字段编码, "operator": "eq/ne/gt/ge/lt/le/in", "value": 值}；
        字段名白名单校验（仅本表 active 字段的 physical_name/code），值全部参数化，禁 SQL 拼接注入。

        count_only（2026-08-14 新增）：True 时发 SELECT COUNT(*) 快速返回行数（int，秒级），
        供预组合数据预览等只需统计的场景，避免全量物化。

        2026-08-08：去除 TOP 1000 截断——多表各取前 N 行且物理序不一致，导致不同表匹配不同批次记录
        （真实根因：表 28↔表 22 截断交集为 0 致 NAME/PRICE 全空、表 24 与表 22 截断批次分裂致 1334 漂移）。
        fetchmany 分批转换防 pyodbc C 层一次性物化全部行（全量行仍驻留 Python 内存供折叠）。
        """
        from apps.modeling.views import DataSourceViewSet, _json_safe
        from apps.modeling.models import Field as MField
        from django.db import connections

        ds = table.data_source
        engine = DataSourceViewSet._ENGINE_MAP.get(ds.db_type)
        if not engine:
            return None

        alias = f'_archive_sync_{ds.id}_{table.id}'
        db_config = {
            'ENGINE': engine,
            'NAME': ds.db_name,
            'HOST': ds.host,
            'PORT': str(ds.port),
            'USER': ds.username,
            'PASSWORD': ds.password,
            'ATOMIC_REQUESTS': False,
            'AUTOCOMMIT': True,
            'TIME_ZONE': None,
            'CONN_MAX_AGE': 0,
            'CONN_HEALTH_CHECKS': False,
            'OPTIONS': {},
        }
        if ds.db_type == 'oracle':
            db_config['OPTIONS'] = {'service_name': ds.db_name}
        elif ds.db_type == 'sqlserver':
            db_config['OPTIONS'] = {
                'driver': 'ODBC Driver 18 for SQL Server',
                'extra_params': 'Encrypt=no',
            }
        connections.databases[alias] = db_config
        try:
            conn = connections[alias]
            conn.ensure_connection()
            schema = table.schema or {'postgresql': 'public', 'sqlserver': 'dbo', 'oracle': '', 'mysql': ''}.get(ds.db_type, '')
            ext_table = table.external_table_name
            order_sql = ''
            if order_by:
                if ds.db_type == 'sqlserver':
                    order_sql = f' ORDER BY [{order_by}]'
                elif ds.db_type == 'oracle':
                    order_sql = f' ORDER BY "{order_by}"'
                elif ds.db_type == 'mysql':
                    order_sql = f' ORDER BY `{order_by}`'
                else:
                    order_sql = f' ORDER BY "{order_by}"'
            where_sql = ''
            where_params = []
            if conditions:
                where_sql, where_params = self._build_conditions_sql(table, conditions, ds.db_type)
            with conn.cursor() as cursor:
                if ds.db_type == 'sqlserver':
                    full_table = f'[{schema}].[{ext_table}]'
                elif ds.db_type == 'oracle':
                    owner = schema.upper() if schema else ''
                    full_table = f'"{owner}"."{ext_table}"' if owner else f'"{ext_table}"'
                elif ds.db_type == 'mysql':
                    full_table = f'`{ext_table}`'
                else:
                    full_table = f'"{schema}"."{ext_table}"'
                if count_only:
                    cursor.execute(f'SELECT COUNT(*) FROM {full_table}{where_sql}', where_params)
                    row = cursor.fetchone()
                    return int(row[0]) if row and row[0] is not None else 0
                if ds.db_type == 'sqlserver':
                    cursor.execute(f'SELECT * FROM {full_table}{where_sql}{order_sql}', where_params)
                elif ds.db_type == 'oracle':
                    cursor.execute(f'SELECT * FROM {full_table}{where_sql}{order_sql}', where_params)
                elif ds.db_type == 'mysql':
                    cursor.execute(f'SELECT * FROM {full_table}{where_sql}{order_sql}', where_params)
                else:
                    cursor.execute(f'SELECT * FROM {full_table}{where_sql}{order_sql}', where_params)
                columns = [desc[0] for desc in cursor.description]
                rows = []
                while True:
                    chunk = cursor.fetchmany(10000)
                    if not chunk:
                        break
                    rows.extend({col: _json_safe(val) for col, val in zip(columns, row)} for row in chunk)
                return rows
        finally:
            connections.databases.pop(alias, None)

    def _split_conditions(self, conditions):
        """按 field_source 拆分条件（2026-08-14）：header 条件应用到头表查询（预组合平铺），
        detail 条件应用到明细表查询；无 field_source 标记的条件视为 detail 侧（存量兼容）。

        背景：预组合的筛选条件可能命中头表字段（如价目表 NAME），原实现无条件拼到明细表查询，
        header 字段不在明细表白名单 → ValueError 整表跳过（条件静默失效根因）。"""
        header_conds, detail_conds = [], []
        for cond in conditions or []:
            if cond.get('field_source') == 'header':
                header_conds.append(cond)
            else:
                detail_conds.append(cond)
        return (header_conds or None), (detail_conds or None)

    def _build_conditions_sql(self, table, conditions, db_type):
        """结构化条件 → 方言化 WHERE 子句（字段白名单校验 + 值参数化，禁拼接注入）。"""
        from apps.modeling.models import Field as MField
        valid_phys = {}
        for f in MField.objects.filter(table=table, status=MField.Status.ACTIVE):
            phys = f.physical_name or f.code
            if phys:
                valid_phys[phys] = phys
                valid_phys.setdefault(f.code, phys)
        if db_type == 'sqlserver':
            def _q(col):
                return f'[{col}]'
        elif db_type == 'oracle':
            def _q(col):
                return f'"{col}"'
        elif db_type == 'mysql':
            def _q(col):
                return f'`{col}`'
        else:
            def _q(col):
                return f'"{col}"'
        clauses = []
        params = []
        for cond in conditions or []:
            col = str(cond.get('field', '')).strip()
            op = str(cond.get('operator', 'eq')).strip()
            val = cond.get('value')
            if col not in valid_phys:
                raise ValueError(f'条件字段 {col} 不在表 {table.name} 字段白名单中')
            quoted = _q(valid_phys[col])
            if op == 'eq':
                clauses.append(f'{quoted} = %s'); params.append(val)
            elif op == 'ne':
                clauses.append(f'{quoted} <> %s'); params.append(val)
            elif op == 'gt':
                clauses.append(f'{quoted} > %s'); params.append(val)
            elif op == 'ge':
                clauses.append(f'{quoted} >= %s'); params.append(val)
            elif op == 'lt':
                clauses.append(f'{quoted} < %s'); params.append(val)
            elif op == 'le':
                clauses.append(f'{quoted} <= %s'); params.append(val)
            elif op == 'in':
                if not isinstance(val, list):
                    raise ValueError(f'条件 {col} 的 in 操作符 value 必须是数组')
                placeholders = ', '.join(['%s'] * len(val))
                clauses.append(f'{quoted} IN ({placeholders})'); params.extend(val)
            elif op == 'starts_with':
                clauses.append(f'{quoted} LIKE %s'); params.append(f'{val}%')
            elif op == 'contains':
                clauses.append(f'{quoted} LIKE %s'); params.append(f'%{val}%')
            else:
                raise ValueError(f'不支持的条件操作符 {op}（支持 eq/ne/gt/ge/lt/le/in/starts_with/contains）')
        if not clauses:
            return '', []
        return ' WHERE ' + ' AND '.join(clauses), params

    def _join_header_rows(self, table, cfg, rows, join_type='left', conditions=None, header_rows=None):
        """预组合平铺（2026-08-11 第三轮）：头表全量拉取，按 header_link_field↔detail_link_field
        JOIN 进明细行。头表字段以 `__hdr__{物理列名}` 前缀并入（与明细字段重名不冲突）；
        头表查询失败或未命中时降级保留纯明细行（不阻断同步）。
        同值多行取排序后最后一条（确定性，与 nested_sources 一致）。

        2026-08-13 Issue 3：join_type='inner' 时无匹配头表的明细行不保留。
        2026-08-14：conditions 透传头表查询（header 侧筛选条件，如价目表 NAME eq 新明码实价）；
        条件命中后头表只拉满足行 → inner 语义天然过滤无匹配明细行（等价用户 SQL INNER JOIN + WHERE）。

        header_rows（2026-08-14 新增）：调用方已拉取头表行时传入复用（顺带可拿头表命中统计），
        避免重复查询；None 时内部照旧自拉。"""
        from apps.modeling.models import Field as MField
        header_table = cfg.header_table
        h_link = cfg.header_link_field
        d_link = cfg.detail_link_field
        h_phys = h_link.physical_name or h_link.code
        d_phys = d_link.physical_name or d_link.code
        t_pk = MField.objects.filter(
            table=header_table, is_primary_key=True, status=MField.Status.ACTIVE).first()
        if header_rows is None:
            header_rows = self._query_external_table(
                header_table, order_by=(t_pk.physical_name or t_pk.code) if t_pk else None,
                conditions=conditions)
        if header_rows is None:
            return rows  # 头表查询失败：降级为纯明细（头字段缺失，同步不阻断）
        hindex = {}
        for hr in header_rows:
            val = hr.get(h_phys)
            if val is None:
                continue
            hindex[str(val)] = hr
        out = []
        for row in rows:
            dv = row.get(d_phys)
            hr = hindex.get(str(dv)) if dv is not None else None
            if hr:
                merged = dict(row)
                for k, v in hr.items():
                    merged[f'__hdr__{k}'] = v
                out.append(merged)
            elif join_type == 'left':
                out.append(row)
            # join_type='inner' 时不保留未匹配头表的行
        return out

    def _build_precombine_filters(self, domain, tables, pk_fields, code_to_physical, match_channels, stats):
        """预组合过滤预扫（2026-08-14）：join_type=inner 的 detail 挂载在同步前预扫 kept_keys
        （满足全部挂载条件的主记录主键值集合），返回 {表id: row_filter} 供主表/直连表 upsert 行级过滤。

        - 每个 inner 挂载：带条件查明细表（+头表按 header 条件 inner 过滤）→ 挂载字段值集合；
        - 挂载字段物理列不在明细/头表（异名挂载，如 FID→MATERIAL_GROUP 物理列在第三张表）→
          桥接查询挂载字段所在表：source_field 值 → 主表主键值（对齐用户 SQL 多表 INNER JOIN 语义）；
        - 多挂载取交集（AND，对齐用户 SQL 多个 INNER JOIN）；
        - kept 空 → warning 该挂载不参与过滤；全部剔除/交集空 → 跳过过滤（warning，防误全量停用）；
        - 仅 upsert 阶段行级过滤（防 stale 复活死循环）：detail 挂载/中转路径天然只命中已过滤的主记录。
        """
        from apps.modeling.models import DetailTableConfig, FieldMapping, Field as MField
        precombine_filters = {}
        if not pk_fields:
            return precombine_filters
        primary_table = next((t for t in tables if t.is_primary), None)
        if not primary_table:
            return precombine_filters
        # 主表主键物理列（各表各自的物理列名）
        pk_phys_by_table = {}
        for pk in pk_fields:
            for tbl_id, phys in list(code_to_physical.get(pk, [])) + list(match_channels.get(pk, [])):
                pk_phys_by_table.setdefault(tbl_id, phys)
        if not pk_phys_by_table.get(primary_table.id):
            return precombine_filters

        kept_sets = []
        for table in tables:
            if not table.data_source:
                continue
            cfg = DetailTableConfig.objects.filter(domain=domain, table=table).first()
            fms = []
            if cfg:
                fms = list(FieldMapping.objects.filter(
                    detail_config=cfg, source_field__status='active',
                ).select_related('source_field', 'target_field', 'detail_config'))
            else:
                fms = list(FieldMapping.objects.filter(
                    source_table=table, relation_type=FieldMapping.RelationType.DETAIL,
                    source_field__status='active',
                ).select_related('source_field', 'target_field', 'detail_config'))
            for fm in fms:
                if fm.join_type != 'inner':
                    continue  # 仅 inner 挂载参与主记录过滤（left 不收敛数据）
                target_code = fm.target_field.code if fm.target_field else None
                if not target_code:
                    continue
                conds = None
                if cfg and cfg.conditions:
                    conds = cfg.conditions
                elif fm.conditions:
                    conds = fm.conditions
                header_conds, detail_conds = self._split_conditions(conds)
                rows = self._query_external_table(table, order_by=None, conditions=detail_conds)
                if rows is None:
                    stats['warnings'].append(
                        f'预组合过滤：{table.name or table.code} 查询失败，该挂载不参与主记录过滤')
                    continue
                if cfg and cfg.header_table_id and cfg.header_link_field_id and cfg.detail_link_field_id:
                    rows = self._join_header_rows(table, cfg, rows, fm.join_type, conditions=header_conds)
                # —— 挂载键值集合：明细行 source_field 物理列行内取值（挂载键=明细行该列值）——
                # 2026-08-14 修复：原实现按 target_code 查 code_to_physical 取 phys_cols，挂载字段
                # code 非 schema code（如 MATERIAL_ID/MATERIAL_GROUP）时解析全空（kept 空 → 过滤
                # 静默失效根因）；改为 source 侧行内取值，天然对齐用户 SQL 的 ON 语义
                # （明细表挂载列值 = 主记录挂载键值）。
                src_phys = fm.source_field.physical_name or fm.source_field.code
                if not src_phys:
                    stats['warnings'].append(
                        f'预组合过滤：{table.name or table.code} 挂载未配置源字段，该挂载不参与主记录过滤')
                    continue
                src_values = set()
                for row in rows:
                    v = row.get(src_phys)
                    if v is None:
                        v = row.get(f'__hdr__{src_phys}')
                    if v is not None:
                        src_values.add(str(v))
                if not src_values:
                    stats['warnings'].append(
                        f'预组合过滤：{table.name or table.code} 挂载 {target_code} 条件未命中任何明细行，'
                        f'该挂载不参与过滤')
                    continue
                # —— 主记录侧挂载键通道 → kept（主键值集合，row_filter 统一按主键比较）——
                # 挂载键值域 == target 表主键值域（同域，如价目明细 MATERIAL_ID ↔ 物料主键）→
                # 直接以明细行值作为 kept；否则桥接 target_field 所在表（{主键值 → 挂载键值}，
                # 如分组头 FID → 物料表 MATERIAL_GROUP），主记录行按主键值查桥接后比较。
                tf = fm.target_field
                tf_phys = tf.physical_name or tf.code
                tf_table = tf.table
                same_domain = False
                if tf_table and tf_table.data_source:
                    tf_pk = MField.objects.filter(
                        table=tf_table, is_primary_key=True, status=MField.Status.ACTIVE).first()
                    if tf_pk and (tf_pk.physical_name or tf_pk.code) == tf_phys:
                        same_domain = True
                if same_domain:
                    kept = src_values
                else:
                    kept = set()
                    if tf_table and tf_table.data_source:
                        tf_pk = MField.objects.filter(
                            table=tf_table, is_primary_key=True, status=MField.Status.ACTIVE).first()
                        if tf_pk:
                            tf_pk_phys = tf_pk.physical_name or tf_pk.code
                            brow = self._query_external_table(tf_table, order_by=None)
                            if brow is not None:
                                for r in brow:
                                    pv = r.get(tf_pk_phys)
                                    tv = r.get(tf_phys)
                                    if pv is None or tv is None:
                                        continue
                                    if str(tv) in src_values:
                                        kept.add(str(pv))
                if kept:
                    kept_sets.append(kept)
                else:
                    stats['warnings'].append(
                        f'预组合过滤：{table.name or table.code} 挂载 {target_code} 条件未命中任何主记录，'
                        f'该挂载不参与过滤')
        if not kept_sets:
            stats['warnings'].append('预组合过滤：所有 inner 挂载条件均未命中，跳过主记录过滤')
            return precombine_filters
        kept = set.intersection(*kept_sets)
        if not kept:
            stats['warnings'].append(
                '预组合过滤：inner 挂载条件交集为空（各条件分别命中但无共同主记录），跳过主记录过滤')
            return precombine_filters
        for tbl_id, phys in pk_phys_by_table.items():
            precombine_filters[tbl_id] = (lambda row, _phys=phys, _kept=kept:
                                          str(row.get(_phys)) in _kept)
        return precombine_filters

    def _sync_detail_rows(self, archive, table, rows, detail_fm, code_to_physical, pk_fields,
                          match_channels, operated_by, stats, matched_ids, change_entries,
                          created_in_this_batch, sync_exclude_codes):
        """子表关系同步（2026-08-08）：明细行全量保留写 ArchiveRecordDetail + 代表行折叠写主表。

        - 归属（2026-08-13 方向修正）：明细行按挂载字段（detail_fm.target_field）值归属主记录，
          同值多记录=一对多全部挂载（替代原按主表主键归属）；
        - 行键：detail_fm.row_key_field 配置优先；未配置 → _detect_unique_column 自动检测唯一列并回填配置；
        - 嵌套属性：target_table=本表的 reference 映射（一级透传），源表属性以 `__nested__{code}` 前缀
          并入明细行（同值多行取排序后最后一条，确定性）；
        - 代表行：display_sort_field 排序（空值垫底）+ 行键 DESC 次级键取首条写主表
          （生效日期最新 + 同日期取行键最大，确定性可复现）；
        - 明细 upsert：source_data 整层替换 + manual_data 保留，merged 有差异才 save；
          批1 明细变更不进 change_entries（防假明细 BUG-2026-0805-01 教训），批2 扩展 ChangeDetail 时统一加；
        - 明细停用清扫：源侧消失的行标 DELETED（安全闸门：无同步错误时执行）。
        """
        from apps.modeling.models import FieldMapping, Field as MField

        schema = archive.schema or []
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in schema}
        source_table_name = table.name or table.code

        # 本表物理列 → schema code（本表映射字段写入用）
        # 2026-08-11 第三轮：预组合平铺时头表物理列同样纳入（__hdr__ 前缀字段按基础列名映射）
        physical_to_schema = {}
        detail_cfg0 = detail_fm.detail_config
        header_table_id = detail_cfg0.header_table_id if (detail_cfg0 and detail_cfg0.header_table_id) else None
        for schema_code, mappings in code_to_physical.items():
            for tbl_id, phys_col in mappings:
                if tbl_id == table.id or tbl_id == header_table_id:
                    physical_to_schema[phys_col] = schema_code

        # 归属键：挂载字段（主表端 target_field，2026-08-13 方向修正：任何键均可挂载，
        # 不再限定主键；同步按挂载字段值匹配主记录，同值多记录=一对多全部挂载）
        # 本表物理列 → 挂载字段 schema code（明细归属主记录 + 代表行 key 构建用）
        # 2026-08-11 第三轮修复：预组合时头表物理列同样纳入（头表字段可作挂载关联键，
        # 平铺行中以其 `__hdr__` 前缀形态被 _record_key_for_row 取值）
        # 2026-08-14：target_code 优先标准字段解析——挂载字段 code 可能非 schema code
        # （如物料表 MATERIAL_ID 的 std=MTL_ID，主记录 data 键是 MTL_ID），用 code 直配会全失配
        tf0 = detail_fm.target_field
        target_code = None
        if tf0:
            sf0 = tf0.standard_field
            target_code = sf0.standard_code if (sf0 and sf0.standard_code) else tf0.code
        target_physical_to_schema = {}
        if target_code:
            for tbl_id, phys in list(code_to_physical.get(target_code, [])) + list(match_channels.get(target_code, [])):
                if tbl_id == table.id or tbl_id == header_table_id:
                    target_physical_to_schema[phys] = target_code
        # 2026-08-14 异名挂载补洞：挂载字段物理列不在本表/头表（如分组头 FID→物料 MATERIAL_GROUP，
        # 物理列在第三张表物料表）时，用 source_field 物理列作行内取值通道——行内值=明细行
        # source_field 值，即用户 SQL 的 ON MTL1.MATERIAL_GROUP=GG.FID 语义（组合体主键=挂载键）。
        if (not target_physical_to_schema and detail_fm.source_field and detail_fm.target_field):
            src_phys = detail_fm.source_field.physical_name or detail_fm.source_field.code
            if src_phys:
                target_physical_to_schema[src_phys] = target_code
        if not target_code or not target_physical_to_schema:
            stats['warnings'].append(
                f'明细子表 {source_table_name}：挂载字段 {target_code or "未配置"} 无本表物理列映射，'
                f'明细行无法归属主记录（请在关系管理重新选择挂载字段）')
            return

        # 行键列：detail_config 优先；未配置自动检测唯一列并回填配置（一次检测全表复用）
        detail_cfg = detail_fm.detail_config
        row_key_field = detail_cfg.row_key_field if (detail_cfg and detail_cfg.row_key_field_id) else detail_fm.row_key_field
        row_key_phys = None
        if row_key_field:
            row_key_phys = row_key_field.physical_name or row_key_field.code
        else:
            row_key_phys = self._detect_unique_column(table, rows)
            if not row_key_phys:
                stats['warnings'].append(
                    f'明细子表 {source_table_name}：未配置行键列且自动检测未找到唯一列，'
                    f'跳过明细同步（可在关系管理手动指定行键列）')
                return
            mf = (MField.objects.filter(table=table, physical_name=row_key_phys).first()
                  or MField.objects.filter(table=table, code=row_key_phys).first())
            if mf:
                if detail_cfg:
                    detail_cfg.row_key_field = mf
                    detail_cfg.save(update_fields=['row_key_field'])
                else:
                    detail_fm.row_key_field = mf
                    detail_fm.save(update_fields=['row_key_field'])
                stats['warnings'].append(
                    f'明细子表 {source_table_name}：行键列自动检测为 {row_key_phys}，已回填关系配置')

        # 代表行排序字段（空值垫底；日期 ISO 字典序=时间序；行键次级键保证确定性）
        display_field = detail_cfg.display_sort_field if (detail_cfg and detail_cfg.display_sort_field_id) else detail_fm.display_sort_field
        display_phys = (display_field.physical_name or display_field.code) if display_field else None
        sort_desc = detail_cfg.display_sort_desc if (detail_cfg and detail_cfg.id) else detail_fm.display_sort_desc
        if display_phys is None:
            stats['warnings'].append(
                f'明细子表 {source_table_name}：未配置代表行排序字段，主表展示字段未更新'
                f'（可在关系管理配置，如 EFFECTIVE_DATE）')

        # 嵌套属性源表：target_table=本表的 reference 映射（一级透传，如 27.NAME/DESCRIPTION 并进 28 行）
        nested_sources = []
        for fm in FieldMapping.objects.filter(
                target_table=table,
                relation_type=FieldMapping.RelationType.REFERENCE,
                source_field__status='active', target_field__status='active',
        ).select_related('source_table', 'source_field', 'target_field'):
            src = fm.source_table
            if not src.data_source:
                continue
            src_phys = fm.source_field.physical_name or fm.source_field.code
            tgt_phys = fm.target_field.physical_name or fm.target_field.code
            if not src_phys or not tgt_phys:
                continue
            # 源表映射到 schema 的字段（仅透传已释放字段）
            src_physical_to_schema = {}
            for schema_code, mappings in code_to_physical.items():
                for tbl_id, phys_col in mappings:
                    if tbl_id == src.id:
                        src_physical_to_schema[phys_col] = schema_code
            t_pk = MField.objects.filter(
                table=src, is_primary_key=True, status=MField.Status.ACTIVE).first()
            src_rows = self._query_external_table(
                src, order_by=(t_pk.physical_name or t_pk.code) if t_pk else None)
            if src_rows is None:
                continue
            tindex = {}
            for srow in src_rows:
                val = srow.get(src_phys)
                if val is None:
                    continue
                tindex[str(val)] = srow  # 同值多行取排序后最后一条（确定性）
            nested_sources.append((tgt_phys, tindex, src_physical_to_schema, src.id, fm.join_type))

        def _rk(v):
            return '' if v is None else str(v)

        # 预加载本档案已有明细（含停用），按 (record_id, row_key) 索引
        existing_details = {}
        for d in ArchiveRecordDetail.objects.filter(
                mapping=detail_fm, record__archive=archive).order_by('id'):
            existing_details[(d.record_id, d.row_key)] = d

        # 预加载该档案全部记录（active 优先），代表行/明细归属匹配用
        # 2026-08-13 方向修正：按挂载字段值索引，同值多记录全保留（一对多挂载）
        # 2026-08-14 桥接：挂载键不在档案 schema（如分组头 FID→物料 MATERIAL_GROUP，主记录
        # data 无该键）时，经 target_field 所在表桥接 {主键值 → 挂载键值} 索引
        schema_codes = {i.get('code') for i in schema}
        bridge_by_pk = None
        if target_code and target_code not in schema_codes and tf0 and tf0.table:
            btable = tf0.table
            b_pk = MField.objects.filter(
                table=btable, is_primary_key=True, status=MField.Status.ACTIVE).first()
            b_phys = tf0.physical_name or tf0.code
            if btable.data_source and b_pk and b_phys:
                b_pk_phys = b_pk.physical_name or b_pk.code
                brow = self._query_external_table(btable, order_by=None)
                if brow is not None:
                    bridge_by_pk = {}
                    for r in brow:
                        pv = r.get(b_pk_phys)
                        tv = r.get(b_phys)
                        if pv is not None and tv is not None:
                            bridge_by_pk[str(pv)] = str(tv)
        existing_records = {}
        for rec in ArchiveRecord.objects.filter(archive=archive).order_by('id'):
            if bridge_by_pk is not None:
                pk_v = None
                for pk in pk_fields:
                    v0 = rec.data.get(pk)
                    if v0 is not None:
                        pk_v = v0
                        break
                val = bridge_by_pk.get(str(pk_v)) if pk_v is not None else None
            else:
                val = rec.data.get(target_code)
            if val is None or str(val) == '':
                continue
            key = str(val)
            lst = existing_records.setdefault(key, [])
            if rec.status == ArchiveRecord.Status.ACTIVE:
                lst.insert(0, rec)
            else:
                lst.append(rec)

        def _record_key_for_row(row):
            # 挂载字段值（含平铺头表 __hdr__ 前缀列）；无值返回 None（无法归属）
            for phys, code in target_physical_to_schema.items():
                if phys in row and row[phys] is not None:
                    return str(row[phys])
                hdr_key = f'__hdr__{phys}'
                if hdr_key in row and row[hdr_key] is not None:
                    return str(row[hdr_key])
            return None

        # 代表行排序（display_sort DESC/ASC + 行键次级键；空值永远垫底）
        sorted_rows = rows
        if display_phys:
            non_null = [r for r in rows if r.get(display_phys) is not None]
            null_rows = [r for r in rows if r.get(display_phys) is None]
            non_null.sort(key=lambda r: (str(r.get(display_phys)), _rk(r.get(row_key_phys))),
                          reverse=sort_desc)
            sorted_rows = non_null + null_rows

        detail_no_change = []
        record_no_change = []
        matched_detail_ids = set()
        blank_rk = 0
        for row in sorted_rows:
            # —— 明细行归属主记录（行内挂载字段物理列构建 key，2026-08-13 一对多）——
            rec_key = _record_key_for_row(row)
            if rec_key is None:
                continue  # 行内无挂载字段值：无法归属
            existing_list = existing_records.get(rec_key, [])
            if not existing_list:
                continue  # 挂载字段值未匹配到任何主记录：跳过

            # —— 明细行数据：本表映射字段 + 嵌套属性透传（__nested__ 前缀独立命名空间）——
            # 2026-08-11 第三轮：__hdr__ 前缀字段按基础列名映射头表物理列（预组合平铺）
            detail_data = {}
            for col_name, value in row.items():
                is_hdr = col_name.startswith('__hdr__')
                base_col = col_name[7:] if is_hdr else col_name
                schema_code = physical_to_schema.get(base_col)
                if not schema_code:
                    continue
                if (table.id, base_col) in sync_exclude_codes and schema_code not in pk_fields:
                    continue
                detail_data[schema_code] = value
            skip_row = False
            for tgt_phys, tindex, src_phys_to_schema, src_id, nested_join_type in nested_sources:
                tgt_val = row.get(tgt_phys)
                if tgt_val is None:
                    if nested_join_type == 'inner':
                        skip_row = True
                    continue
                srow = tindex.get(str(tgt_val))
                if not srow:
                    if nested_join_type == 'inner':
                        skip_row = True
                    continue
                for sc, sv in srow.items():
                    sc_code = src_phys_to_schema.get(sc)
                    if not sc_code or sc_code in pk_fields:
                        continue
                    if (src_id, sc) in sync_exclude_codes:
                        continue
                    detail_data[f'__nested__{sc_code}'] = sv
            if skip_row:
                continue
            if not detail_data:
                continue  # 无任何档案字段的明细行不落库

            rk = _rk(row.get(row_key_phys))
            if not rk:
                blank_rk += 1
                continue  # 行键为空：无法唯一定位明细（行键列配置错误，防 unique_together 冲突）
            # 2026-08-13 一对多：同挂载字段值的所有主记录都挂该明细行
            for existing in existing_list:
                existing_detail = existing_details.get((existing.id, rk))
                if existing_detail:
                    matched_detail_ids.add(existing_detail.id)
                    # 源删自动停用的明细行源端重现 → 自动复活
                    if existing_detail.status == ArchiveRecordDetail.Status.DELETED:
                        existing_detail.status = ArchiveRecordDetail.Status.ACTIVE
                    # 整层替换：明细行全部数据来自本表（无他表合并），源侧删字段即消失
                    existing_detail.source_data = detail_data
                    merged, lineage = _merge_record_data(existing_detail, schema)
                    old_data = existing_detail.data or {}
                    if old_data != merged:
                        existing_detail.data = merged
                        existing_detail.lineage = lineage
                        existing_detail.save()
                        stats['details_updated'] += 1
                    else:
                        existing_detail.lineage = lineage
                        detail_no_change.append(existing_detail)
                else:
                    detail = ArchiveRecordDetail(
                        record=existing, mapping=detail_fm, row_key=rk,
                        source_data=detail_data, manual_data={},
                    )
                    merged, lineage = _merge_record_data(detail, schema)
                    detail.data = merged
                    detail.lineage = lineage
                    detail.save()
                    existing_details[(existing.id, rk)] = detail
                    matched_detail_ids.add(detail.id)
                    stats['details_created'] += 1

        if blank_rk:
            stats['warnings'].append(
                f'明细子表 {source_table_name}：{blank_rk} 行行键为空已跳过（请检查行键列配置）')

        # 收尾：无数据差异的明细统一批量落库（含复活状态/血缘）
        if detail_no_change:
            ArchiveRecordDetail.objects.bulk_update(
                detail_no_change,
                ['source_data', 'manual_data', 'lineage', 'status'],
                batch_size=2000,
            )

        # —— 代表行写主表（按挂载字段值分组：每组排序首行 = 默认价；对齐第133轮方向锁定语义）——
        # 2026-08-13 一对多：同挂载字段值的所有主记录共享代表行数据
        # 复用 _write_dimension_row 公共写入逻辑：本表非空映射字段 → source_data 合并 → 版本+1 + 变更明细
        if display_phys is not None and sorted_rows:
            seen_keys = set()
            for rep_row in sorted_rows:
                rep_key = _record_key_for_row(rep_row)
                if rep_key is None or rep_key in seen_keys:
                    continue
                seen_keys.add(rep_key)
                for rep_existing in existing_records.get(rep_key, []):
                    self._write_dimension_row(
                        rep_existing, rep_row, physical_to_schema, schema, field_name_map,
                        source_table_name, (rep_key,), operated_by, stats, matched_ids,
                        change_entries, created_in_this_batch, record_no_change,
                    )

        # 收尾：代表行变更的无差异记录批量落库
        if record_no_change:
            ArchiveRecord.objects.bulk_update(
                record_no_change,
                ['source_data', 'manual_data', 'lineage', 'sync_status', 'status'],
                batch_size=2000,
            )

        # —— 明细停用清扫：源侧消失的明细行标 DELETED（安全闸门：无错误时执行）——
        if not stats['errors']:
            candidate_ids = set(ArchiveRecordDetail.objects.filter(
                mapping=detail_fm, status=ArchiveRecordDetail.Status.ACTIVE,
            ).values_list('id', flat=True))
            stale_ids = sorted(candidate_ids - matched_detail_ids)
            for i in range(0, len(stale_ids), 500):
                ArchiveRecordDetail.objects.filter(id__in=stale_ids[i:i + 500]).update(
                    status=ArchiveRecordDetail.Status.DELETED)
            stats['details_deactivated'] += len(stale_ids)

        # 批2：明细聚合变更日志（统计级，不逐行；防假明细 BUG-2026-0805-01）
        created = stats.get('details_created', 0)
        updated = stats.get('details_updated', 0)
        deactivated = stats.get('details_deactivated', 0)
        if created > 0 or updated > 0 or deactivated > 0:
            change_entries.append({
                'record_id': None,
                'record_key': f'{source_table_name} 明细',
                'change_type': ArchiveChangeDetail.ChangeType.DETAIL_SYNC,
                'field_changes': [{'detail_stats': {
                    'created': created, 'updated': updated, 'deactivated': deactivated,
                }}],
                'version_before': None,
                'version_after': None,
                'detail_group': None,
                'detail_row_key': '',
            })

    def _detect_unique_column(self, table, rows):
        """全量行逐列统计唯一性：无空值且 COUNT(DISTINCT)==总行数 = 候选唯一列。
        优先已标主键（is_primary_key）列，其次按物理列稳定序取第一个。
        覆盖反例：FID 标主键但仅 14,883/239,504 唯一（标主键≠行唯一，须实测）。"""
        from apps.modeling.models import Field as MField
        if not rows:
            return None
        total = len(rows)
        cols = list(rows[0].keys())
        counts = {c: set() for c in cols}
        for r in rows:
            for c in cols:
                s = counts[c]
                if s is None:
                    continue
                v = r.get(c)
                if v is None or v == '':
                    counts[c] = None  # 空值 → 不可用作行键
                else:
                    s.add(str(v))
        candidates = [c for c in cols if counts[c] is not None and len(counts[c]) == total]
        if not candidates:
            return None
        pk_names = set()
        for f in MField.objects.filter(table=table, is_primary_key=True, status=MField.Status.ACTIVE):
            pk_names.add(f.physical_name or f.code)
        for c in candidates:
            if c in pk_names:
                return c
        return candidates[0]

    def _upsert_records_from_rows(self, archive, table, rows, code_to_physical, schema_type_map, pk_fields, operated_by, stats, matched_ids=None, change_entries=None, created_in_this_batch=None, sync_exclude_codes=None, match_channels=None, is_primary_table=False, row_filter=None):
        """将查询结果行写入档案（双层存储，换底重合并）：

        - 该表映射到的字段值直接写入 source_data 底层（零比对，无保护/覆盖分支）；
        - data = _merge_record_data 写时合并物化（archive 字段 manual_data 优先，人工值天然保留）；
        - 合并结果与旧 data 有差异才 version+1 并留版本快照（change_summary.source_refreshed）；
        - 停用记录也参与匹配：源删自动停用（sync_status='stale'）的记录源端重现时自动复活；
          手工停用的记录只更新数据保持停用，不再重建重复记录；
        - 每条变更（新增/修改/复活）追加到 change_entries，供收尾落变更日志批次；
        - 匹配通道（match_channels）：组合字段非主成员列（外键）仅用于构建记录 key，不写入档案；
        - 确定性折叠：同 key 多行保留排序后的最后一行（1:n 取首条），并计入 cardinality_fold_count；
        - 非主表匹配不到已有记录时跳过不创建（外键引用不是独立实体，防数据爆炸）；
        - 预组合过滤（2026-08-14）：row_filter 非空时仅处理满足条件的行（不满足的行不进 seen_keys/
          不 upsert/不创建，防 stale 记录被源行命中自动复活形成过滤死循环）。
        """
        schema = archive.schema or []
        if matched_ids is None:
            matched_ids = set()
        if change_entries is None:
            change_entries = []
        if created_in_this_batch is None:
            created_in_this_batch = set()
        if sync_exclude_codes is None:
            sync_exclude_codes = set()
        if match_channels is None:
            match_channels = {}
        # 无数据差异的记录收集批量落库（防全量模式下每轮 20 万次逐条 UPDATE）
        no_change_updates = []
        # 字段 code → 中文名（变更明细展示用）
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in schema}

        # 构建该表的物理字段 code → schema code 的反向映射
        physical_to_schema = {}
        for schema_code, mappings in code_to_physical.items():
            for tbl_id, phys_col in mappings:
                if tbl_id == table.id:
                    physical_to_schema[phys_col] = schema_code

        # 匹配通道：本表中映射到主键 schema code 的物理列（不写入，仅构建 key 用）
        pk_physical_to_schema = {}
        for pk_code, entries in match_channels.items():
            for tbl_id, phys in entries:
                if tbl_id == table.id:
                    pk_physical_to_schema[phys] = pk_code

        # 预加载该档案全部记录（含停用），用主键值建索引；同主键时 active 优先
        existing_records = {}
        for rec in ArchiveRecord.objects.filter(archive=archive).order_by('id'):
            key = tuple(str(rec.data.get(pk, '')) for pk in pk_fields)
            if not any(k for k in key):  # 至少需一个主键值非空
                continue
            prev = existing_records.get(key)
            if prev is not None and prev.status == ArchiveRecord.Status.ACTIVE:
                continue
            if prev is None or rec.status == ArchiveRecord.Status.ACTIVE:
                existing_records[key] = rec

        source_table_name = table.name or table.code
        now_iso = timezone.now().isoformat()

        # 预组合过滤（2026-08-14）：行级过滤必须在 upsert 阶段（防 stale 复活死循环——
        # 不满足条件的行不进 seen_keys、不 upsert、不进 matched_ids，下一轮停用清扫统一标 stale）
        if row_filter is not None:
            rows = [row for row in rows if row_filter(row)]

        # 预解析 + 按主键折叠：同 key 多行保留排序后的最后一行（确定性取首条）
        parsed_rows = []
        fold_keys = set()  # 出现 >1 行的 key（1:n 发散检查）
        seen_keys = {}
        for row in rows:
            record_data = {}
            pk_match_values = {}
            for col_name, value in row.items():
                # 先解析 schema code
                schema_code = physical_to_schema.get(col_name)
                if not schema_code:
                    # 匹配通道：组合字段非主成员列（外键）仅用于构建记录 key，不写入
                    pk_code = pk_physical_to_schema.get(col_name)
                    if pk_code:
                        pk_match_values[pk_code] = value
                        continue
                    # 同名兜底仅限主键列（记录匹配必需）：未映射给本表的其他同名列
                    # 不得写入，否则他表同名空列会偷渡清空已有值造成假变更（BUG-2026-0805-01）
                    if col_name in pk_fields and col_name in schema_type_map:
                        schema_code = col_name
                if not schema_code:
                    continue
                # 组合字段的非主字段成员不写入档案（只用于一致性检查）
                # 但主键字段必须保留用于记录匹配，即使它在排除集合中
                if (table.id, col_name) in sync_exclude_codes and schema_code not in pk_fields:
                    continue
                record_data[schema_code] = value

            if not record_data and not pk_match_values:
                continue

            # 用主键值匹配已有记录；无主键值的源行不进档案（无法匹配，避免每轮刷新重建）
            key = tuple(str(record_data.get(pk, pk_match_values.get(pk, ''))) for pk in pk_fields)
            if not any(k for k in key):
                continue
            if key in seen_keys:
                fold_keys.add(key)
            seen_keys[key] = (record_data, pk_match_values)

        if fold_keys and stats is not None:
            stats['cardinality_fold_count'] = stats.get('cardinality_fold_count', 0) + len(fold_keys)
            table_label = table.name or table.code
            stats['cardinality_warnings'] = stats.get('cardinality_warnings', [])
            stats['cardinality_warnings'].append(
                f'表 {table_label}: {len(fold_keys)} 个主键对应多条源行，已按确定性排序取首条折叠'
            )

        for key, (record_data, pk_match_values) in seen_keys.items():
            existing = existing_records.get(key)

            if existing:
                matched_ids.add(existing.id)
                # 源删自动停用的记录源端重现 → 自动复活；手工停用（非 stale）保持停用
                reactivated = False
                ver_before = existing.version  # v18 版本映射：变更前版本号
                if existing.status == ArchiveRecord.Status.DELETED and existing.sync_status == 'stale':
                    existing.status = ArchiveRecord.Status.ACTIVE
                    stats['records_reactivated'] += 1
                    reactivated = True
                # 换底：该表映射到的字段直接写入底层（零比对）
                existing.source_data = {**(existing.source_data or {}), **record_data}
                merged, lineage = _merge_record_data(existing, schema)
                # 本表字段的 sync 血缘补 source_table
                for code in record_data:
                    entry = lineage.get(code)
                    if entry and entry.get('source') == 'sync' and entry.get('source_table') != source_table_name:
                        lineage[code] = {**entry, 'source_table': source_table_name}

                old_data = existing.data or {}
                changed_codes = sorted(
                    c for c in set(list(merged.keys()) + list(old_data.keys()))
                    if old_data.get(c) != merged.get(c)
                )
                existing.lineage = lineage
                existing.sync_status = 'synced'
                if changed_codes:
                    existing.data = merged
                    existing.updated_by = operated_by
                    existing.version += 1
                    existing.save()
                    ss, _ = ArchiveSchemaSnapshot.objects.get_or_create(
                        archive=archive,
                        schema_version=archive.schema_version,
                        defaults={'schema': archive.schema},
                    )
                    ArchiveRecordVersion.objects.create(
                        record=existing,
                        version=existing.version,
                        data=existing.data,
                        schema_version_ref=ss,
                        schema=None,
                        operated_by=operated_by,
                        operation_type=ArchiveRecordVersion.OperationType.UPDATE,
                        change_summary={
                            'source_refreshed': changed_codes,
                            'changed_fields': [
                                {'field': c, 'old': old_data.get(c), 'new': merged.get(c)}
                                for c in changed_codes
                            ],
                        },
                    )
                    # 只有当记录不是本轮刚创建时才计入「修改」（后续表合并数据到刚创建的记录不算修改）
                    if existing.id not in created_in_this_batch:
                        stats['records_updated'] += 1
                else:
                    # 合并结果无变化：仅落底层/血缘（含可能的复活状态），不动版本号；批量收集收尾统一更新
                    no_change_updates.append(existing)
                # 变更日志：复活优先于修改；字段级旧值→新值
                # 本轮刚创建的记录被后续表合并时不重复记 UPDATED（创建时已记 CREATED，
                # 与 records_updated 统计口径一致，防首次全量 20 万条重复明细爆炸）
                if reactivated or (changed_codes and existing.id not in created_in_this_batch):
                    change_entries.append({
                        'record_id': existing.id,
                        'record_key': '/'.join(k for k in key),
                        'change_type': (ArchiveChangeDetail.ChangeType.REACTIVATED if reactivated
                                        else ArchiveChangeDetail.ChangeType.UPDATED),
                        'field_changes': [
                            {'field': c, 'name': field_name_map.get(c, c),
                             'old': old_data.get(c), 'new': merged.get(c)}
                            for c in changed_codes
                        ],
                        # v18 版本映射：无数据差异（仅复活）时版本号不变，两者相等
                        'version_before': ver_before,
                        'version_after': existing.version,
                    })
            elif is_primary_table:
                # 创建新记录：底层=源数据，覆盖层空，data=合并结果
                # 仅主表可创建新实体；非主表匹配不到即跳过（外键引用不是独立实体，防数据爆炸）
                record = ArchiveRecord(
                    archive=archive,
                    source_data=record_data,
                    manual_data={},
                    sync_status='synced',
                    created_by=operated_by,
                    updated_by=operated_by,
                )
                merged, lineage = _merge_record_data(record, schema)
                for code in record_data:
                    entry = lineage.get(code)
                    if entry and entry.get('source') == 'sync':
                        lineage[code] = {**entry, 'source_table': source_table_name}
                record.data = merged
                record.lineage = lineage
                record.save()
                ss, _ = ArchiveSchemaSnapshot.objects.get_or_create(
                    archive=archive,
                    schema_version=archive.schema_version,
                    defaults={'schema': archive.schema},
                )
                ArchiveRecordVersion.objects.create(
                    record=record,
                    version=1,
                    data=record.data,
                    schema_version_ref=ss,
                    schema=None,
                    operated_by=operated_by,
                    operation_type=ArchiveRecordVersion.OperationType.CREATE,
                    change_summary={
                        'action': '源同步创建记录',
                        'changed_fields': [
                            {'field': c, 'old': None, 'new': v}
                            for c, v in (record.data or {}).items()
                            if v not in (None, '')
                        ],
                    },
                )
                # 加入索引以便后续表能匹配到这条记录
                existing_records[key] = record
                matched_ids.add(record.id)
                created_in_this_batch.add(record.id)  # 标记为本轮刚创建的记录
                stats['records_created'] += 1
                # 变更日志：新增只记记录级，不展开全部字段值
                change_entries.append({
                    'record_id': record.id,
                    'record_key': '/'.join(k for k in key),
                    'change_type': ArchiveChangeDetail.ChangeType.CREATED,
                    'field_changes': [],
                    # v18 版本映射：新建无变更前版本
                    'version_before': None,
                    'version_after': 1,
                })

        # 收尾：无数据差异的记录统一批量落库（含复活状态）
        if no_change_updates:
            ArchiveRecord.objects.bulk_update(
                no_change_updates,
                ['source_data', 'manual_data', 'lineage', 'sync_status', 'status'],
                batch_size=2000,
            )

    def _upsert_dimension_via_mapping(self, archive, table, rows, code_to_physical, schema_type_map,
                                      pk_fields, match_channels, operated_by, stats, matched_ids=None,
                                      change_entries=None, created_in_this_batch=None):
        """维度表经 FieldMapping 中转匹配主记录（如价目表 NAME 经明细 FID 关联物料）。

        - FieldMapping(source_table=本表)：本表 source_field 值 → target 表 target_field 值（引用关系）；
        - target 行中映射到主键的物理列（code_to_physical + match_channels）→ 构建记录 key → 匹配主记录；
        - 本表映射字段（如 NAME）折叠写入匹配到的记录（确定性排序取首条）；
        - 仅处理一级中转；目标行无法匹配主记录时跳过（外键引用非独立实体）。
        """
        from apps.modeling.models import FieldMapping, Field as MField

        schema = archive.schema or []
        if matched_ids is None:
            matched_ids = set()
        if change_entries is None:
            change_entries = []
        if created_in_this_batch is None:
            created_in_this_batch = set()
        # 无数据差异的记录收集批量落库（与 _upsert_records_from_rows 同口径）
        no_change_updates = []
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in schema}
        source_table_name = table.name or table.code

        # 本表 物理列 → schema code（本表映射字段写入用）
        physical_to_schema = {}
        for schema_code, mappings in code_to_physical.items():
            for tbl_id, phys_col in mappings:
                if tbl_id == table.id:
                    physical_to_schema[phys_col] = schema_code

        fms = FieldMapping.objects.filter(
            source_table=table, source_field__status='active'
        ).select_related('source_field', 'target_table', 'target_field')

        # 预加载该档案全部记录（含停用），用主键值建索引；同主键时 active 优先（一级/多级共用）
        existing_records = {}
        for rec in ArchiveRecord.objects.filter(archive=archive).order_by('id'):
            key = tuple(str(rec.data.get(pk, '')) for pk in pk_fields)
            if not any(k for k in key):
                continue
            prev = existing_records.get(key)
            if prev is not None and prev.status == ArchiveRecord.Status.ACTIVE:
                continue
            if prev is None or rec.status == ArchiveRecord.Status.ACTIVE:
                existing_records[key] = rec

        # —— 多级中转（2026-08-10）：本表无 source 映射时，沿 FieldMapping 无向等值图 BFS
        # 找「可映射档案主键的表」路径逐级透传匹配（如分组多语言 26.FID→25.FID→22.MATERIAL_GROUP→
        # 22.MATERIAL_ID→MTL_ID 两级链；原引擎仅支持一级，整表静默跳过）——
        if not fms:
            chain, target_pk_phys = self._build_mapping_chain(
                table, pk_fields, code_to_physical, match_channels)
            if chain is not None:
                self._apply_mapping_chain(
                    archive, table, rows, chain, target_pk_phys, pk_fields, code_to_physical,
                    schema_type_map, existing_records, operated_by, stats, matched_ids,
                    change_entries, created_in_this_batch,
                )
            return

        # 各中转目标表：查询行，构建 {目标字段值: [目标行...]} 索引（2026-08-10 改：收集全部匹配行供展开传播）
        for fm in fms:
            target = fm.target_table
            if not target.data_source:
                continue
            target_phys = fm.target_field.physical_name or fm.target_field.code
            t_pk_field = MField.objects.filter(
                table=target, is_primary_key=True, status=MField.Status.ACTIVE
            ).first()
            t_order_by = t_pk_field.physical_name or t_pk_field.code if t_pk_field else None
            # 2026-08-13 批2：普通关联支持筛选条件（仅 reference 生效，detail 行为不变）
            conds = None
            if fm.relation_type == FieldMapping.RelationType.REFERENCE and fm.conditions:
                conds = fm.conditions
            trows = self._query_external_table(target, order_by=t_order_by, conditions=conds)
            if trows is None:
                continue
            tindex = {}
            for trow in trows:
                val = trow.get(target_phys)
                if val is None:
                    continue
                # 同值多行全部收集：展开传播（组→组内全部物料）；目标键唯一时等价折叠
                tindex.setdefault(str(val), []).append(trow)

            # target 行中映射到主键的物理列 → schema code
            t_pk_physical_to_schema = {}
            for pk in pk_fields:
                for tbl_id, phys in list(code_to_physical.get(pk, [])) + list(match_channels.get(pk, [])):
                    if tbl_id == target.id:
                        t_pk_physical_to_schema[phys] = pk
            if not t_pk_physical_to_schema:
                continue

            # 本表行处理
            src_phys = fm.source_field.physical_name or fm.source_field.code
            for row in rows:
                src_val = row.get(src_phys)
                if src_val is None:
                    continue
                trows_match = tindex.get(str(src_val))
                if not trows_match:
                    if fm.join_type == 'inner':
                        continue
                    # join_type='left'：无匹配目标行时保留源行（映射字段为空）
                    trows_match = []
                # 展开：对每个匹配目标行写（每组属性传播到组内全部物料）；
                # 目标键唯一时仅一行，等价原折叠取首条语义
                for trow in trows_match:
                    # 从目标行提取主键值
                    key_parts = []
                    for pk in pk_fields:
                        pk_val = None
                        for phys, code in t_pk_physical_to_schema.items():
                            if code == pk and phys in trow:
                                pk_val = trow[phys]
                                break
                        key_parts.append('' if pk_val is None else str(pk_val))
                    key = tuple(key_parts)
                    if not any(key):
                        continue
                    existing = existing_records.get(key)
                    if not existing:
                        continue  # 目标行无法匹配主记录：跳过（外键引用非独立实体）
                    # 折叠写入公共逻辑（与多级中转共用）
                    self._write_dimension_row(
                        existing, row, physical_to_schema, schema, field_name_map,
                        source_table_name, key, operated_by, stats, matched_ids,
                        change_entries, created_in_this_batch, no_change_updates,
                    )

        # 收尾：无数据差异的记录统一批量落库
        if no_change_updates:
            ArchiveRecord.objects.bulk_update(
                no_change_updates,
                ['source_data', 'manual_data', 'lineage', 'sync_status', 'status'],
                batch_size=2000,
            )

    def _build_mapping_chain(self, table, pk_fields, code_to_physical, match_channels):
        """BFS 沿 FieldMapping 无向等值图找「可映射档案主键的表」路径（多级中转）。

        FieldMapping 语义为等值关联（source_field 值 ↔ target_field 值），双向可用；
        从本表出发逐级透传，直到某表的物理列能映射到档案主键（code_to_physical/match_channels）。

        返回 (chain, target_pk_phys) 或 (None, None)：
        - chain = [(查值列, 下一表id, 下一表列), ...]：首条查值列是本表列，其余是上一级表的列；
        - target_pk_phys = {终表物理列: pk schema code}。
        """
        from apps.modeling.models import FieldMapping as FM
        from collections import deque

        # 无向等值边：本表列 ↔ 邻表列
        edges = []
        for fm in FM.objects.filter(source_field__status='active', target_field__status='active'):
            ca = fm.source_field.physical_name or fm.source_field.code or fm.source_field.name
            cb = fm.target_field.physical_name or fm.target_field.code or fm.target_field.name
            edges.append((fm.source_table_id, ca, fm.target_table_id, cb))
        adj = {}
        for a, ca, b, cb in edges:
            adj.setdefault(a, []).append((b, ca, cb))
            adj.setdefault(b, []).append((a, cb, ca))

        # 可映射档案主键的表及其主键物理列
        pk_mapped = {}
        for pk in pk_fields:
            for tbl_id, phys in list(code_to_physical.get(pk, [])) + list(match_channels.get(pk, [])):
                pk_mapped.setdefault(tbl_id, {})[phys] = pk
        if not pk_mapped:
            return None, None

        start = table.id
        if start in pk_mapped:
            return [], pk_mapped[start]
        queue = deque([(start, [])])
        visited = {start}
        while queue:
            cur, path = queue.popleft()
            if len(path) >= 5:
                continue  # 深度上限，防异常长链
            for nxt, cur_col, nxt_col in adj.get(cur, []):
                if nxt in visited:
                    continue
                visited.add(nxt)
                new_path = path + [(cur_col, nxt, nxt_col)]
                if nxt in pk_mapped:
                    return new_path, pk_mapped[nxt]
                queue.append((nxt, new_path))
        return None, None

    def _apply_mapping_chain(self, archive, table, rows, chain, target_pk_phys, pk_fields,
                             code_to_physical, schema_type_map, existing_records, operated_by,
                             stats, matched_ids, change_entries, created_in_this_batch):
        """多级 FieldMapping 中转：沿等值链逐级行透传匹配主记录，折叠写入本表字段。

        链上每级构建 {下一表列值: 行} 索引（同值多行取排序后最后一条，与一级同口径）；
        本表行取值 → 逐级命中中间表行并提取下一跳列值 → 终表行提取主键 → 匹配档案记录。
        """
        from apps.modeling.models import Table as MTable, Field as MField

        schema = archive.schema or []
        field_name_map = {i.get('code'): i.get('name') or i.get('code') for i in schema}
        source_table_name = table.name or table.code

        # 本表 物理列 → schema code（本表映射字段写入用，与一级同口径）
        physical_to_schema = {}
        for schema_code, mappings in code_to_physical.items():
            for tbl_id, phys_col in mappings:
                if tbl_id == table.id:
                    physical_to_schema[phys_col] = schema_code

        # 逐级构建索引：中间表单行（透传用），终表收集全部匹配行（展开传播）
        indexes = []
        for from_col, to_tid, to_col in chain:
            to_table = MTable.objects.get(id=to_tid)
            if not to_table.data_source:
                return
            t_pk = MField.objects.filter(
                table=to_table, is_primary_key=True, status=MField.Status.ACTIVE).first()
            t_order_by = (t_pk.physical_name or t_pk.code) if t_pk else None
            trows = self._query_external_table(to_table, order_by=t_order_by)
            if trows is None:
                return
            is_last = (to_tid == chain[-1][1])
            tindex = {}
            for trow in trows:
                val = trow.get(to_col)
                if val is None:
                    continue
                if is_last:
                    # 终表：同值多行全部收集（组→组内全部物料展开传播）
                    tindex.setdefault(str(val), []).append(trow)
                else:
                    tindex[str(val)] = trow  # 中间表透传取排序后最后一条
            indexes.append((from_col, tindex, is_last))

        no_change_updates = []
        for row in rows:
            v = row.get(chain[0][0])
            if v is None:
                continue
            # 逐级透传：中间表单行命中 → 取下一跳列值（等值链值域可能不同列）
            ok = True
            for i in range(len(indexes) - 1):
                trow = indexes[i][1].get(str(v))
                if not trow:
                    ok = False
                    break
                v = trow.get(chain[i + 1][0])
                if v is None:
                    ok = False
                    break
            if not ok:
                continue
            # 终表：展开写全部匹配行（目标键唯一时仅一行，等价折叠）
            final_index = indexes[-1][1]
            for trow in final_index.get(str(v), []):
                # 终表行提取主键值
                key_parts = []
                for pk in pk_fields:
                    pk_val = None
                    for phys, code in target_pk_phys.items():
                        if code == pk and phys in trow:
                            pk_val = trow[phys]
                            break
                    key_parts.append('' if pk_val is None else str(pk_val))
                key = tuple(key_parts)
                if not any(key):
                    continue
                existing = existing_records.get(key)
                if not existing:
                    continue  # 终表行无法匹配主记录：跳过（外键引用非独立实体）
                self._write_dimension_row(
                    existing, row, physical_to_schema, schema, field_name_map,
                    source_table_name, key, operated_by, stats, matched_ids,
                    change_entries, created_in_this_batch, no_change_updates,
                )

        if no_change_updates:
            ArchiveRecord.objects.bulk_update(
                no_change_updates,
                ['source_data', 'manual_data', 'lineage', 'sync_status', 'status'],
                batch_size=2000,
            )

    def _write_dimension_row(self, existing, row, physical_to_schema, schema, field_name_map,
                             source_table_name, key, operated_by, stats, matched_ids,
                             change_entries, created_in_this_batch, no_change_updates):
        """维度行折叠写入公共逻辑（一级/多级中转共用，防复制分叉）：

        本表行非空映射字段 → source_data 合并 → 计算变更 → 版本+1 + 变更明细（或批量无变化落库）。
        """
        # 本表映射字段折叠写入
        record_data = {}
        for col_name, value in row.items():
            schema_code = physical_to_schema.get(col_name)
            if not schema_code:
                continue
            if value is None:
                continue  # 维度字段空值不写（避免清空已有值）
            record_data[schema_code] = value
        if not record_data:
            return
        matched_ids.add(existing.id)
        existing.source_data = {**(existing.source_data or {}), **record_data}
        merged, lineage = _merge_record_data(existing, schema)
        for code in record_data:
            entry = lineage.get(code)
            if entry and entry.get('source') == 'sync' and entry.get('source_table') != source_table_name:
                lineage[code] = {**entry, 'source_table': source_table_name}
        old_data = existing.data or {}
        changed_codes = sorted(
            c for c in set(list(merged.keys()) + list(old_data.keys()))
            if old_data.get(c) != merged.get(c)
        )
        ver_before = existing.version
        existing.lineage = lineage
        existing.sync_status = 'synced'
        if changed_codes:
            existing.data = merged
            existing.updated_by = operated_by
            existing.version += 1
            existing.save()
            # 查找 archive（_write_dimension_row 通过 schema 参数传入了 archive.schema）
            _archive_for_ss = existing.archive
            ss, _ = ArchiveSchemaSnapshot.objects.get_or_create(
                archive=_archive_for_ss,
                schema_version=_archive_for_ss.schema_version,
                defaults={'schema': schema},
            )
            ArchiveRecordVersion.objects.create(
                record=existing,
                version=existing.version,
                data=existing.data,
                schema_version_ref=ss,
                schema=None,
                operated_by=operated_by,
                operation_type=ArchiveRecordVersion.OperationType.UPDATE,
                change_summary={
                    'source_refreshed': changed_codes,
                    'changed_fields': [
                        {'field': c, 'old': old_data.get(c), 'new': merged.get(c)}
                        for c in changed_codes
                    ],
                },
            )
            if existing.id not in created_in_this_batch:
                stats['records_updated'] += 1
        else:
            no_change_updates.append(existing)
        if changed_codes and existing.id not in created_in_this_batch:
            change_entries.append({
                'record_id': existing.id,
                'record_key': '/'.join(k for k in key),
                'change_type': ArchiveChangeDetail.ChangeType.UPDATED,
                'field_changes': [
                    {'field': c, 'name': field_name_map.get(c, c),
                     'old': old_data.get(c), 'new': merged.get(c)}
                    for c in changed_codes
                ],
                'version_before': ver_before,
                'version_after': existing.version,
            })

    def _data_matches(self, existing_data, new_data):
        """简单判断两条记录是否代表同一实体（比较非空字段的交集）"""
        if not existing_data or not new_data:
            return False
        common_keys = set(existing_data.keys()) & set(new_data.keys())
        if not common_keys:
            return False
        match_count = sum(1 for k in common_keys if str(existing_data.get(k, '')) == str(new_data.get(k, '')))
        # 超过 80% 的公共字段相同视为同一记录
        return match_count / len(common_keys) >= 0.8


def refresh_archive_data(archive, operated_by='system'):
    """数据刷新（供 refresh-data 端点 / 定时线程 / management command 复用）：

    仅执行源数据整层拉取（换底重合并）+ 计算字段批量重算 + SYNC 操作日志；
    不重生成 schema、不 bump schema_version。
    """
    import logging
    logger = logging.getLogger(__name__)
    domain = archive.domain
    schema_type_map = {item['code']: item['type'] for item in (archive.schema or []) if item.get('code')}
    viewset = ArchiveViewSet()
    stats = {'records_created': 0, 'records_updated': 0, 'tables_synced': 0, 'errors': [], 'warnings': []}
    try:
        stats = viewset._sync_data_from_sources(archive, domain, schema_type_map, operated_by)
    except Exception as e:
        logger.error(f'档案 {archive.id} 数据刷新失败: {e}')
        stats['errors'].append(str(e))
    if domain:
        try:
            from apps.modeling.computed_service import batch_recalculate
            stats['computed_recalculated'] = batch_recalculate(domain.id)
        except Exception as e:
            logger.warning(f'档案 {archive.id} 计算字段重算失败: {e}')
            stats['computed_recalculated'] = {'error': str(e)}
    ArchiveOperationLog.objects.create(
        archive=archive,
        operator=operated_by,
        operation_type=ArchiveOperationLog.OperationType.SYNC,
        change_summary={'action': '源数据刷新（按主字段从源表拉取并合并）', 'refresh_stats': stats},
    )
    return stats


class ArchiveRecordViewSet(viewsets.ModelViewSet):
    """档案记录 API"""
    queryset = ArchiveRecord.objects.select_related('archive', 'archive__domain').all()
    filterset_fields = ['archive', 'status', 'sync_status']

    def create(self, request, *args, **kwargs):
        """主数据记录统一由业务系统同步产生，禁止档案端人工新增"""
        return Response(
            {'detail': '主数据记录统一由业务系统同步产生，不允许在档案端人工新增'},
            status=status.HTTP_403_FORBIDDEN,
        )

    def get_serializer_class(self):
        if self.action == 'list':
            return ArchiveRecordListSerializer
        elif self.action == 'create':
            return ArchiveRecordCreateSerializer
        elif self.action in ('update', 'partial_update'):
            return ArchiveRecordUpdateSerializer
        return ArchiveRecordDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        archive_id = self.request.query_params.get('archive')
        if archive_id:
            qs = qs.filter(archive_id=archive_id)
        # 关键字搜索：按业务数据内容模糊匹配（JSON 文本化后 icontains，SQLite 兼容）
        search = (self.request.query_params.get('search') or '').strip()
        if search:
            from django.db.models.functions import Cast
            from django.db.models import TextField
            qs = qs.annotate(_data_text=Cast('data', TextField())).filter(_data_text__icontains=search)
        # 改过（未同步）的记录置顶，其余按更新时间倒序
        qs = qs.annotate(
            _sync_rank=Case(
                When(sync_status='synced', then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).order_by('_sync_rank', '-updated_at')
        return qs

    def perform_destroy(self, instance):
        with transaction.atomic():
            old_data = instance.data
            old_status = instance.get_status_display()
            instance.version += 1
            instance.status = ArchiveRecord.Status.DELETED
            instance.save()
            # 记录删除版本（变更内容记全：动作说明 + 状态变化 + 字段终值快照数）
            ArchiveRecordVersion.objects.create(
                record=instance,
                version=instance.version,
                data=old_data,
                schema=instance.archive.schema,
                operated_by='system',
                operation_type=ArchiveRecordVersion.OperationType.DELETE,
                change_summary={
                    'action': '删除记录（软删除，状态置为已停用，数据快照保留可回滚）',
                    'changed_fields': [
                        {'field': '状态', 'old': old_status, 'new': '已停用'},
                    ],
                    'snapshot_field_count': len(old_data or {}),
                },
            )
            # 记录操作日志
            ArchiveOperationLog.objects.create(
                archive=instance.archive,
                record=instance,
                operator='system',
                operation_type=ArchiveOperationLog.OperationType.DELETE,
                change_summary={'action': '删除记录（软删除）'},
            )

    @action(detail=True, methods=['get'], url_path='versions')
    def list_versions(self, request, pk=None):
        """查看版本历史"""
        record = self.get_object()
        versions = ArchiveRecordVersion.objects.filter(record=record)
        page = self.paginate_queryset(versions)
        if page is not None:
            serializer = VersionSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = VersionSerializer(versions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='versions/compare')
    def compare_versions(self, request, pk=None):
        """版本差异对比"""
        record = self.get_object()
        v1 = request.query_params.get('v1')
        v2 = request.query_params.get('v2')
        if not v1 or not v2:
            return Response({'error': '需要提供 v1 和 v2 参数'}, status=status.HTTP_400_BAD_REQUEST)

        version_1 = get_object_or_404(ArchiveRecordVersion, record=record, version=v1)
        version_2 = get_object_or_404(ArchiveRecordVersion, record=record, version=v2)

        data_1 = version_1.data or {}
        data_2 = version_2.data or {}
        diff = []
        all_keys = set(list(data_1.keys()) + list(data_2.keys()))
        for key in sorted(all_keys):
            if data_1.get(key) != data_2.get(key):
                diff.append({
                    'field': key,
                    'old_value': data_1.get(key),
                    'new_value': data_2.get(key),
                })
        return Response({'version_1': int(v1), 'version_2': int(v2), 'diff': diff})

    @action(detail=True, methods=['post'], url_path='rollback')
    def rollback(self, request, pk=None):
        """回滚到指定版本（v18：统一执行器按 ownership 分层写回，
        修复旧实现只写合并层、下次合并/刷新后回滚效果静默消失的隐性 Bug）"""
        record = self.get_object()
        serializer = RollbackSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        target_version = serializer.validated_data['target_version']
        operated_by = serializer.validated_data['operated_by']

        version_snapshot = get_object_or_404(ArchiveRecordVersion, record=record, version=target_version)

        # 快照全字段作目标值，经统一执行器分层写回（source→source_data，archive→manual_data/回落）
        _execute_field_rollback(
            record, dict(version_snapshot.data or {}), operated_by,
            action_text=f'回滚至 v{target_version}（同步状态置为未同步）',
        )
        return Response(ArchiveRecordDetailSerializer(record).data)

    @action(detail=True, methods=['get'], url_path='details')
    def list_detail_rows(self, request, pk=None):
        """获取记录的全部明细子表行"""
        record = self.get_object()
        details = record.details.select_related('mapping').all()
        from .serializers import ArchiveRecordDetailRowSerializer
        serializer = ArchiveRecordDetailRowSerializer(details, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='rollback-to-change')
    def rollback_to_change(self, request, pk=None):
        """按时间点回滚（v18：直接恢复明细对应的版本快照，不再从 field_changes 反推目标值）。

        POST /records/{id}/rollback-to-change/  body: {target_detail_id: int, operated_by: str}
        """
        record = self.get_object()
        target_detail_id = request.data.get('target_detail_id')
        operated_by = request.data.get('operated_by', 'system')

        if not target_detail_id:
            return Response({'error': '必须提供 target_detail_id'}, status=status.HTTP_400_BAD_REQUEST)

        # 确认目标变更明细属于该记录
        target_detail = get_object_or_404(ArchiveChangeDetail, pk=target_detail_id)
        if target_detail.record_id != record.id:
            return Response({'error': '目标变更明细不属于该记录'}, status=status.HTTP_400_BAD_REQUEST)

        # 存量历史明细无版本映射（v18 引入），无法定位时点快照
        if target_detail.version_after is None:
            return Response({'error': '该历史变更缺少版本映射，无法按时间点回滚，请使用版本回滚'},
                            status=status.HTTP_400_BAD_REQUEST)
        if record.version == target_detail.version_after:
            return Response({'error': '该时点即当前状态，无需回滚'}, status=status.HTTP_400_BAD_REQUEST)

        snapshot = ArchiveRecordVersion.objects.filter(
            record=record, version=target_detail.version_after).first()
        if not snapshot:
            return Response({'error': '未找到该时点的版本快照'}, status=status.HTTP_400_BAD_REQUEST)

        result = _execute_field_rollback(
            record, dict(snapshot.data or {}), operated_by,
            action_text=f'回滚到变更明细 #{target_detail_id} 时点（恢复 v{target_detail.version_after} 快照）',
        )
        return Response(result)

    @action(detail=True, methods=['post'], url_path='pin')
    def pin_version(self, request, pk=None):
        """定版当前版本"""
        record = self.get_object()
        operated_by = request.data.get('operated_by', 'system')
        note = request.data.get('note', '')

        # 获取当前版本的快照
        current_version = get_object_or_404(
            ArchiveRecordVersion, record=record, version=record.version
        )
        if current_version.is_pinned:
            return Response({'error': '该版本已定版'}, status=status.HTTP_400_BAD_REQUEST)

        current_version.is_pinned = True
        current_version.pinned_at = timezone.now()
        current_version.pinned_by = operated_by
        current_version.pin_note = note
        current_version.save(update_fields=['is_pinned', 'pinned_at', 'pinned_by', 'pin_note'])

        ArchiveOperationLog.objects.create(
            archive=record.archive,
            record=record,
            operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.PIN,
            change_summary={'action': f'定版 v{record.version}（锁定当前版本快照）', 'version': record.version, 'note': note},
        )
        return Response(VersionSerializer(current_version).data)


class SyncLogViewSet(viewsets.ReadOnlyModelViewSet):
    """同步日志 API（只读）"""
    queryset = ArchiveSyncLog.objects.select_related('archive').all()
    serializer_class = SyncLogSerializer
    filterset_fields = ['archive', 'status']
    ordering = ['-started_at']


class OperationLogViewSet(viewsets.ReadOnlyModelViewSet):
    """操作日志 API（只读）"""
    queryset = ArchiveOperationLog.objects.select_related('archive').all()
    serializer_class = OperationLogSerializer
    filterset_fields = ['archive', 'operation_type', 'operator']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        archive_id = self.request.query_params.get('archive')
        if archive_id:
            qs = qs.filter(archive_id=archive_id)
        return qs


class RecordVersionViewSet(viewsets.ReadOnlyModelViewSet):
    """全局记录版本 API（版本管理页）：跨记录查询版本快照 + 定版/取消定版"""
    queryset = ArchiveRecordVersion.objects.select_related('record', 'record__archive').all()
    serializer_class = GlobalVersionSerializer
    ordering = ['-id']

    def get_queryset(self):
        qs = super().get_queryset().order_by('-id')
        params = self.request.query_params
        if params.get('archive'):
            qs = qs.filter(record__archive_id=params['archive'])
        if params.get('record'):
            qs = qs.filter(record_id=params['record'])
        if params.get('operation_type'):
            qs = qs.filter(operation_type=params['operation_type'])
        if params.get('is_pinned') in ('true', 'false'):
            qs = qs.filter(is_pinned=params['is_pinned'] == 'true')
        if params.get('operated_by'):
            qs = qs.filter(operated_by__icontains=params['operated_by'])
        return qs

    @action(detail=True, methods=['post'], url_path='pin')
    def pin(self, request, pk=None):
        """定版指定版本快照"""
        version = self.get_object()
        if version.is_pinned:
            return Response({'error': '该版本已定版'}, status=status.HTTP_400_BAD_REQUEST)
        operated_by = request.data.get('operated_by', 'system')
        version.is_pinned = True
        version.pinned_at = timezone.now()
        version.pinned_by = operated_by
        version.pin_note = request.data.get('note', '')
        version.save(update_fields=['is_pinned', 'pinned_at', 'pinned_by', 'pin_note'])
        ArchiveOperationLog.objects.create(
            archive=version.record.archive,
            record=version.record,
            operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.PIN,
            change_summary={'action': f'定版 v{version.version}（锁定版本快照）', 'version': version.version, 'note': version.pin_note},
        )
        return Response(GlobalVersionSerializer(version).data)

    @action(detail=True, methods=['post'], url_path='unpin')
    def unpin(self, request, pk=None):
        """取消定版"""
        version = self.get_object()
        if not version.is_pinned:
            return Response({'error': '该版本未定版'}, status=status.HTTP_400_BAD_REQUEST)
        operated_by = request.data.get('operated_by', 'system')
        version.is_pinned = False
        version.pinned_at = None
        version.pinned_by = ''
        version.pin_note = ''
        version.save(update_fields=['is_pinned', 'pinned_at', 'pinned_by', 'pin_note'])
        ArchiveOperationLog.objects.create(
            archive=version.record.archive,
            record=version.record,
            operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.UNPIN,
            change_summary={'action': f'取消定版 v{version.version}', 'version': version.version},
        )
        return Response(GlobalVersionSerializer(version).data)


class ChangeBatchViewSet(viewsets.ReadOnlyModelViewSet):
    """数据变更批次 API（只读 + 整批回滚）"""
    queryset = ArchiveChangeBatch.objects.select_related('archive').all()
    serializer_class = ChangeBatchSerializer
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        if params.get('archive'):
            qs = qs.filter(archive_id=params['archive'])
        if params.get('change_source'):
            qs = qs.filter(change_source=params['change_source'])
        return qs

    @action(detail=False, methods=['post'], url_path='start-manual')
    def start_manual(self, request):
        """开启人工批次（v18 攒批保存）：前端保存时先开批次，随后逐条
        PUT /records/{id}/（带 change_batch_id）将明细全部攒入本批；保存即批次封口。
        POST /change-batches/start-manual/  body: {archive: int, operated_by: str}
        """
        archive_id = request.data.get('archive')
        operated_by = request.data.get('operated_by', 'system')
        if not archive_id:
            return Response({'error': '必须提供 archive 参数'}, status=status.HTTP_400_BAD_REQUEST)
        archive = get_object_or_404(Archive, pk=archive_id)
        batch = ArchiveChangeBatch.objects.create(
            archive=archive,
            change_source=ArchiveChangeBatch.ChangeSource.MANUAL,
            operator=operated_by,
            stats={},
        )
        return Response(ChangeBatchSerializer(batch).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='rollback')
    def rollback_batch(self, request, pk=None):
        """整批撤销（v18，应对源侧批量刷错事故）：将本批影响的记录逐条恢复到本批之前的状态。

        该批之后又被编辑过的记录跳过并列出（防静默覆盖后续人工修改）；
        记录已删除/存量明细无版本映射的跳过计数。
        POST /change-batches/{id}/rollback/  body: {operated_by: str}
        """
        batch = self.get_object()
        operated_by = request.data.get('operated_by', 'system')

        details = list(batch.details.select_related('record'))
        if not details:
            return Response({'error': '该批次无变更明细'}, status=status.HTTP_400_BAD_REQUEST)

        rollback_batch_obj = ArchiveChangeBatch.objects.create(
            archive=batch.archive,
            change_source=ArchiveChangeBatch.ChangeSource.MANUAL,
            operator=operated_by,
            stats={},
        )
        rolled_back_records = 0
        skipped_edited = []
        skipped_deleted = 0
        skipped_legacy = 0
        for d in details:
            if d.record_id is None:
                skipped_deleted += 1
                continue
            if d.version_before is None:
                skipped_legacy += 1
                continue
            record = d.record
            # 该批之后又被改过（审核标记不算数据变更，不算）→ 跳过并列出
            later_edited = ArchiveChangeDetail.objects.filter(record=record, id__gt=d.id)\
                .exclude(change_type__in=['created', 'reviewed', 'ignored']).exists()
            if later_edited:
                skipped_edited.append({'record_key': d.record_key, 'record_label': d.record_label})
                continue
            snapshot = ArchiveRecordVersion.objects.filter(
                record=record, version=d.version_before).first()
            if snapshot is None:
                skipped_legacy += 1
                continue
            result = _execute_field_rollback(
                record, dict(snapshot.data or {}), operated_by,
                action_text=f'撤销批次 #{batch.id}（{batch.get_change_source_display()}）',
                change_batch=rollback_batch_obj,
            )
            if result['rolled_back_fields'] > 0:
                rolled_back_records += 1

        rollback_batch_obj.stats = {
            'records_rolled_back': rolled_back_records,
            'skipped_edited': len(skipped_edited),
            'skipped_deleted': skipped_deleted,
            'skipped_legacy': skipped_legacy,
            'source_batch_id': batch.id,
        }
        rollback_batch_obj.save(update_fields=['stats'])
        return Response({
            'rolled_back_records': rolled_back_records,
            'skipped_edited': skipped_edited,
            'skipped_deleted': skipped_deleted,
            'skipped_legacy': skipped_legacy,
            'batch_id': rollback_batch_obj.id,
        })


class ChangeDetailViewSet(viewsets.ReadOnlyModelViewSet):
    """数据变更明细 API（只读）"""
    queryset = ArchiveChangeDetail.objects.select_related('batch', 'archive').all()
    serializer_class = ChangeDetailSerializer
    ordering = ['-id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        if params.get('archive'):
            qs = qs.filter(archive_id=params['archive'])
        if params.get('batch'):
            qs = qs.filter(batch_id=params['batch'])
        if params.get('record'):
            qs = qs.filter(record_id=params['record'])
        if params.get('change_type'):
            qs = qs.filter(change_type=params['change_type'])
        if params.get('change_source'):
            qs = qs.filter(batch__change_source=params['change_source'])
        if params.get('record_key'):
            qs = qs.filter(record_key__icontains=params['record_key'])
        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """导出单个档案的全部变更日志 Excel（Sheet1 批次汇总 + Sheet2 变更明细）。

        必须指定 archive 参数；明细上限 50000 行（超出只导最新，末行提示）。
        """
        from urllib.parse import quote
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        archive_id = request.query_params.get('archive')
        if not archive_id:
            return Response({'error': '导出必须指定 archive 参数（单个档案）'}, status=status.HTTP_400_BAD_REQUEST)
        archive = get_object_or_404(Archive, pk=archive_id)

        MAX_ROWS = 50000
        source_label = dict(ArchiveChangeBatch.ChangeSource.choices)
        type_label = dict(ArchiveChangeDetail.ChangeType.choices)
        header_font = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical='top')

        wb = openpyxl.Workbook()

        # Sheet1 批次汇总
        ws1 = wb.active
        ws1.title = '批次汇总'
        ws1.append(['批次号', '时间', '来源', '操作人', '新增', '修改', '停用', '复活', '明细数'])
        for c in ws1[1]:
            c.font = header_font
        batches = (ArchiveChangeBatch.objects.filter(archive=archive)
                   .annotate(detail_cnt=Count('details')).order_by('-created_at'))
        for b in batches:
            s = b.stats or {}
            ws1.append([
                b.id,
                timezone.localtime(b.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                source_label.get(b.change_source, b.change_source),
                b.operator,
                s.get('records_created', 0), s.get('records_updated', 0),
                s.get('records_deactivated', 0), s.get('records_reactivated', 0),
                b.detail_cnt,
            ])

        # Sheet2 变更明细
        ws2 = wb.create_sheet('变更明细')
        ws2.append(['明细ID', '批次号', '时间', '来源', '操作人', '类型', '记录标识', '字段变更（旧值 → 新值）'])
        for c in ws2[1]:
            c.font = header_font
        details = (ArchiveChangeDetail.objects.filter(archive=archive)
                   .select_related('batch').order_by('-id')[:MAX_ROWS + 1])
        truncated = False
        for i, d in enumerate(details):
            if i >= MAX_ROWS:
                truncated = True
                break
            changes_text = '\n'.join(
                f"{fc.get('name') or fc.get('field')}：{fc.get('old')} → {fc.get('new')}"
                for fc in (d.field_changes or [])
            )
            ws2.append([
                d.id, d.batch_id,
                timezone.localtime(d.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                source_label.get(d.batch.change_source, d.batch.change_source),
                d.batch.operator,
                type_label.get(d.change_type, d.change_type),
                d.record_key,
                changes_text,
            ])
            ws2.cell(row=ws2.max_row, column=8).alignment = wrap
        if truncated:
            ws2.append(['', '', '', '', '', '', '', f'明细超过 {MAX_ROWS} 行，仅导出最新 {MAX_ROWS} 条'])

        for ws, widths in ((ws1, [10, 20, 12, 16, 8, 8, 8, 8, 10]),
                           (ws2, [10, 10, 20, 12, 16, 10, 22, 60])):
            for idx, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = w
            ws.freeze_panes = 'A2'

        filename = f"变更日志_{archive.name}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        wb.save(resp)
        return resp

    @action(detail=True, methods=['post'], url_path='rollback')
    def rollback_change(self, request, pk=None):
        """单条变更明细回滚（v18 语义：恢复到本条变更之前的状态，本条之后的变更会一并撤销）。

        POST /change-details/{id}/rollback/  body: {operated_by: str}
        存量历史明细（version_before 为 NULL）降级回旧字段级 old 值恢复逻辑兼容。
        """
        detail = self.get_object()
        operated_by = request.data.get('operated_by', 'system')

        if not detail.record_id:
            return Response({'error': '该记录已被删除，无法回滚'}, status=status.HTTP_400_BAD_REQUEST)
        if detail.change_type in ('created', 'rollback'):
            return Response({'error': f'类型为「{detail.get_change_type_display()}」的变更不支持回滚'},
                            status=status.HTTP_400_BAD_REQUEST)

        record = detail.record

        # v18 新语义：恢复 version_before 快照（本条变更之前的状态）
        if detail.version_before is not None:
            snapshot = ArchiveRecordVersion.objects.filter(
                record=record, version=detail.version_before).first()
            if snapshot is not None:
                result = _execute_field_rollback(
                    record, dict(snapshot.data or {}), operated_by,
                    action_text=f'回滚变更明细 #{detail.id}（恢复到本条变更前 v{detail.version_before}）',
                )
                return Response(result)
            # 快照缺失（异常数据）→ 落入下方存量兼容路径

        # 存量兼容路径：字段级恢复到 old 值
        rollback_fields = {fc['field']: fc['old']
                          for fc in (detail.field_changes or [])
                          if fc.get('field') and fc['field'] != '状态'}
        if not rollback_fields:
            return Response({'error': '该变更无可回滚的字段'}, status=status.HTTP_400_BAD_REQUEST)

        result = _execute_field_rollback(
            record, rollback_fields, operated_by,
            action_text=f'回滚变更明细 #{detail.id}',
        )
        return Response(result)


def _execute_field_rollback(record, target_fields, operated_by, action_text='', change_batch=None):
    """共用回滚执行器：将记录指定字段恢复到目标值（v18：支持共享批次 + 版本映射写入）。

    target_fields: {field_code: target_value}
    change_batch: 传入时变更留痕明细挂入该批次（批次级回滚共用）；否则单独建批次
    返回：{rolled_back_fields, batch_id, new_version}
    """
    from .serializers import _record_pk_key, _composite_label_codes, _build_record_label

    archive = record.archive
    schema = archive.schema or []
    schema_map = {item.get('code'): item for item in schema}

    old_data = dict(record.data or {})
    source_layer = dict(record.source_data or {})
    manual_layer = dict(record.manual_data or {})

    # 按字段 ownership 分层写入
    actual_changes = []  # 实际发生变化的字段
    for code, target_val in target_fields.items():
        current_val = old_data.get(code)
        if current_val == target_val:
            continue  # 已经是目标值，跳过
        item = schema_map.get(code, {})
        ownership = item.get('ownership') or 'archive'
        if ownership == 'source':
            # 源系统维护字段：直接写 source_data，清 manual 遗留
            source_layer[code] = target_val
            manual_layer.pop(code, None)
        else:
            # 档案维护字段：目标值 == 源层则回落，否则写 manual
            if code in source_layer and source_layer.get(code) == target_val:
                manual_layer.pop(code, None)
            else:
                manual_layer[code] = target_val
        name = item.get('name') or code
        actual_changes.append({'field': code, 'name': name, 'old': current_val, 'new': target_val})

    if not actual_changes:
        return {'rolled_back_fields': 0, 'batch_id': None, 'new_version': record.version,
                'message': '所有字段已是目标值，无需回滚'}

    # 写入双层 + 合并物化
    record.source_data = source_layer
    record.manual_data = manual_layer
    merged, lineage = _merge_record_data(record, schema)
    record.data = merged
    record.lineage = lineage
    ver_before = record.version  # v18 版本映射：变更前版本号
    record.version += 1
    record.updated_by = operated_by
    record.sync_status = 'unsynced'

    with transaction.atomic():
        record.save()

        # 版本快照
        change_summary = {'action': action_text, 'changed_fields': actual_changes}
        ss, _ = ArchiveSchemaSnapshot.objects.get_or_create(
            archive=archive,
            schema_version=archive.schema_version,
            defaults={'schema': schema},
        )
        ArchiveRecordVersion.objects.create(
            record=record,
            version=record.version,
            data=record.data,
            schema_version_ref=ss,
            schema=None,
            operated_by=operated_by,
            operation_type=ArchiveRecordVersion.OperationType.ROLLBACK,
            change_summary=change_summary,
        )
        # 操作日志
        ArchiveOperationLog.objects.create(
            archive=archive, record=record, operator=operated_by,
            operation_type=ArchiveOperationLog.OperationType.ROLLBACK,
            change_summary=change_summary,
        )
        # 变更日志留痕（change_type=rollback）；批次级回滚时挂入共享批次
        batch = change_batch or ArchiveChangeBatch.objects.create(
            archive=archive,
            change_source=ArchiveChangeBatch.ChangeSource.MANUAL,
            operator=operated_by,
            stats={'records_rolled_back': 1},
        )
        ArchiveChangeDetail.objects.create(
            batch=batch, archive=archive, record=record,
            record_key=_record_pk_key(record),
            record_label=_build_record_label(_composite_label_codes(archive.domain), record.data),
            change_type=ArchiveChangeDetail.ChangeType.ROLLBACK,
            field_changes=actual_changes,
            version_before=ver_before,
            version_after=record.version,
        )

    return {
        'rolled_back_fields': len(actual_changes),
        'batch_id': batch.id,
        'new_version': record.version,
        'changes': actual_changes,
    }


class ConsistencyIssueViewSet(viewsets.ReadOnlyModelViewSet):
    """一致性差异记录（只读列表 + 批量标记）；不回写任何源表。"""
    serializer_class = ConsistencyIssueSerializer

    def get_queryset(self):
        qs = ConsistencyIssue.objects.select_related('archive').prefetch_related('value_history').order_by('-id')
        p = self.request.query_params
        if p.get('archive'):
            qs = qs.filter(archive_id=p['archive'])
        if p.get('status'):
            qs = qs.filter(status=p['status'])
        if p.get('field_code'):
            qs = qs.filter(field_code=p['field_code'])
        if p.get('record_key'):
            qs = qs.filter(record_key__icontains=p['record_key'])
        if p.get('check_type'):
            qs = qs.filter(check_type=p['check_type'])
        return qs

    @action(detail=False, methods=['post'], url_path='batch-review')
    def batch_review(self, request):
        """批量标记：reviewed（已审核）/ ignored（忽略）/ reopen（重新打开）。

        标记操作写入变更日志批次（change_source='consistency'）+明细（差异快照），
        返回统计结果；已消失（resolved）的差异不可再标记（reopen 除外）。
        """
        ids = request.data.get('ids') or []
        act = request.data.get('action')
        note = (request.data.get('note') or '')[:500]
        operated_by = request.data.get('operated_by', 'system')
        if not ids or act not in ('reviewed', 'ignored', 'reopen'):
            return Response({'error': '参数错误：ids 不能为空，action 必须是 reviewed/ignored/reopen'},
                            status=status.HTTP_400_BAD_REQUEST)
        issues = list(ConsistencyIssue.objects.filter(id__in=ids))
        if not issues:
            return Response({'error': '未找到差异记录'}, status=status.HTTP_400_BAD_REQUEST)

        target = {'reviewed': ConsistencyIssue.Status.REVIEWED,
                  'ignored': ConsistencyIssue.Status.IGNORED,
                  'reopen': ConsistencyIssue.Status.OPEN}[act]
        now = timezone.now()
        updated, by_archive = 0, {}
        for i in issues:
            if act != 'reopen' and i.status == ConsistencyIssue.Status.RESOLVED:
                continue  # 已消失的差异无需审核
            if i.status == target:
                continue
            old_status = i.get_status_display()
            i.status = target
            if act == 'reopen':
                i.review_note, i.reviewed_by, i.reviewed_at = '', '', None
            else:
                i.review_note, i.reviewed_by, i.reviewed_at = note, operated_by, now
            i.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at'])
            updated += 1
            by_archive.setdefault(i.archive_id, []).append((i, old_status))

        # 变更日志：每档案一个批次，零变更不建批次
        type_map = {'reviewed': ArchiveChangeDetail.ChangeType.REVIEWED,
                    'ignored': ArchiveChangeDetail.ChangeType.IGNORED,
                    'reopen': ArchiveChangeDetail.ChangeType.UPDATED}
        batch_ids = []
        for archive_id, items in by_archive.items():
            batch = ArchiveChangeBatch.objects.create(
                archive_id=archive_id,
                change_source=ArchiveChangeBatch.ChangeSource.CONSISTENCY,
                operator=operated_by,
                stats={'action': act, 'issues_marked': len(items), 'note': note},
            )
            ArchiveChangeDetail.objects.bulk_create([
                ArchiveChangeDetail(
                    batch=batch, archive_id=archive_id, record_id=i.record_id,
                    record_key=i.record_key[:200],
                    change_type=type_map[act],
                    field_changes=[{'field': i.field_code, 'name': i.field_name or i.field_code,
                                    'old': i.member_value, 'new': i.primary_value}],
                ) for (i, _old) in items
            ])
            batch_ids.append(batch.id)
        return Response({'updated': updated, 'skipped': len(issues) - updated,
                         'action': act, 'batch_ids': batch_ids})


class ConsistencyCheckRuleViewSet(viewsets.ModelViewSet):
    """一致性检查规则失效管理：列表/创建/更新/删除。

    用户可以将特定检查规则失效，失效后该规则产生的差异不计入统计。
    """
    serializer_class = ConsistencyCheckRuleSerializer

    def get_queryset(self):
        qs = ConsistencyCheckRule.objects.select_related('archive').order_by('-disabled_at')
        p = self.request.query_params
        if p.get('archive'):
            qs = qs.filter(archive_id=p['archive'])
        if p.get('check_type'):
            qs = qs.filter(check_type=p['check_type'])
        if p.get('disabled') is not None:
            qs = qs.filter(disabled=p['disabled'] == 'true')
        return qs

    @action(detail=True, methods=['post'], url_path='toggle')
    def toggle(self, request, pk=None):
        """切换规则失效/启用状态"""
        rule = self.get_object()
        rule.disabled = not rule.disabled
        if rule.disabled:
            rule.disabled_by = request.data.get('operated_by', 'system')
            rule.disabled_reason = (request.data.get('reason') or '')[:500]
        else:
            rule.disabled_by = ''
            rule.disabled_reason = ''
        rule.save(update_fields=['disabled', 'disabled_by', 'disabled_reason'])
        return Response(ConsistencyCheckRuleSerializer(rule).data)

    @action(detail=False, methods=['post'], url_path='disable')
    def disable_rule(self, request):
        """将指定规则失效。参数：archive, check_type, field_code, member_source, reason, operated_by"""
        archive_id = request.data.get('archive')
        check_type = request.data.get('check_type')
        field_code = request.data.get('field_code', '')
        member_source = request.data.get('member_source', '')
        if not archive_id or not check_type:
            return Response({'error': '参数错误：archive 和 check_type 必填'},
                            status=status.HTTP_400_BAD_REQUEST)
        rule, created = ConsistencyCheckRule.objects.get_or_create(
            archive_id=archive_id, check_type=check_type,
            field_code=field_code, member_source=member_source,
            defaults={
                'disabled': True,
                'disabled_by': request.data.get('operated_by', 'system'),
                'disabled_reason': (request.data.get('reason') or '')[:500],
            }
        )
        if not created:
            rule.disabled = True
            rule.disabled_by = request.data.get('operated_by', 'system')
            rule.disabled_reason = (request.data.get('reason') or '')[:500]
            rule.save(update_fields=['disabled', 'disabled_by', 'disabled_reason'])
        return Response(ConsistencyCheckRuleSerializer(rule).data)

    @action(detail=False, methods=['post'], url_path='enable')
    def enable_rule(self, request):
        """恢复指定规则。参数：archive, check_type, field_code, member_source"""
        archive_id = request.data.get('archive')
        check_type = request.data.get('check_type')
        field_code = request.data.get('field_code', '')
        member_source = request.data.get('member_source', '')
        try:
            rule = ConsistencyCheckRule.objects.get(
                archive_id=archive_id, check_type=check_type,
                field_code=field_code, member_source=member_source)
            rule.disabled = False
            rule.disabled_by = ''
            rule.disabled_reason = ''
            rule.save(update_fields=['disabled', 'disabled_by', 'disabled_reason'])
            return Response(ConsistencyCheckRuleSerializer(rule).data)
        except ConsistencyCheckRule.DoesNotExist:
            return Response({'error': '未找到该规则'}, status=status.HTTP_404_NOT_FOUND)


def _match_condition(value, operator, target):
    """在 Python 层判断单个筛选条件是否满足"""
    if value is None:
        return False
    try:
        if operator == 'eq':
            return str(value) == str(target)
        if operator == 'ne':
            return str(value) != str(target)
        if operator == 'contains':
            return str(target) in str(value)
        if operator in ('gt', 'lt'):
            try:
                fv, ft = float(value), float(target)
            except (TypeError, ValueError):
                fv, ft = str(value), str(target)
            return fv > ft if operator == 'gt' else fv < ft
    except Exception:
        return False
    return True


class ArchiveApiViewSet(viewsets.ModelViewSet):
    """数据服务API 管理"""
    queryset = ArchiveApi.objects.select_related('archive', 'archive__domain').all()
    serializer_class = ArchiveApiSerializer
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        archive_id = self.request.query_params.get('archive')
        if archive_id:
            qs = qs.filter(archive_id=archive_id)
        status_val = self.request.query_params.get('status')
        if status_val:
            qs = qs.filter(status=status_val)
        return qs

    @action(detail=True, methods=['get'], url_path='data')
    def data(self, request, pk=None):
        """返回该 API 暴露的字段定义 + 启用数据（按筛选条件过滤）"""
        api_obj = self.get_object()
        archive = api_obj.archive
        full_schema = archive.schema or []
        exposed = api_obj.exposed_fields or []

        # 字段定义：若 exposed 为空则全部，否则取子集（保持 schema 顺序）
        if exposed:
            exposed_set = set(exposed)
            schema = [f for f in full_schema if f.get('code') in exposed_set]
        else:
            schema = full_schema
        field_codes = [f.get('code') for f in schema]

        # 启用记录 + 筛选
        conditions = api_obj.filter_conditions or []
        records = []
        qs = ArchiveRecord.objects.filter(archive=archive, status=ArchiveRecord.Status.ACTIVE)
        for rec in qs:
            data = rec.data or {}
            # 应用筛选条件（AND）
            ok = True
            for cond in conditions:
                fld = cond.get('field')
                op = cond.get('operator', 'eq')
                target = cond.get('value')
                if not _match_condition(data.get(fld), op, target):
                    ok = False
                    break
            if not ok:
                continue
            # 仅保留暴露字段
            row = {code: data.get(code) for code in field_codes}
            row['__id'] = rec.id
            records.append(row)

        return Response({
            'schema': schema,
            'records': records,
            'auth_roles': api_obj.auth_roles or [],
            'filter_conditions': conditions,
            'name': api_obj.name,
            'path': api_obj.path,
            'status': api_obj.status,
        })

    @action(detail=True, methods=['get'], url_path='docs')
    def docs(self, request, pk=None):
        """接口文档管理端预览（与对外 /api/open/{slug}/docs/ 同构，无需密钥）"""
        from .open_api_gateway import build_docs
        api_obj = self.get_object()
        if not api_obj.slug:
            return Response({'detail': '该接口尚未生成对外路径（slug）'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(build_docs(api_obj))


# ===== 域变更统计（问题7：域概览页）=====
from rest_framework.decorators import api_view
from django.db.models import Max, Q
from apps.modeling.models import Domain


# ===== API 密钥管理（v19，REQ-005）=====


class ApiKeyViewSet(viewsets.ModelViewSet):
    """API 密钥管理：创建（明文仅返回一次）/编辑/轮换/吊销/调用日志"""
    queryset = ApiKey.objects.prefetch_related('grants', 'grants__api').all()
    serializer_class = ApiKeySerializer
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        status_val = self.request.query_params.get('status')
        if status_val:
            qs = qs.filter(status=status_val)
        return qs

    @staticmethod
    def _sync_grants(api_key, grants_payload):
        """全量重建授权关系（新建/编辑共用）；操作范围不得超出 API 自身 allowed_operations"""
        from .open_api_auth import OPERATIONS
        if grants_payload is None:
            return
        api_key.grants.all().delete()
        for g in grants_payload:
            api_obj = ArchiveApi.objects.filter(id=g.get('api')).first()
            if api_obj is None:
                continue
            ops = [op for op in (g.get('allowed_operations') or ['read']) if op in OPERATIONS]
            api_ops = api_obj.allowed_operations or ['read']
            ops = [op for op in ops if op in api_ops] or ['read']
            ApiKeyGrant.objects.create(api_key=api_key, api=api_obj, allowed_operations=ops)

    def create(self, request, *args, **kwargs):
        from . import open_api_auth as auth
        plain = auth.generate_api_key()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            api_key = serializer.save(
                key_prefix=auth.key_prefix(plain),
                key_hash=auth.hash_api_key(plain),
            )
            self._sync_grants(api_key, request.data.get('grants'))
        result = ApiKeySerializer(api_key).data
        # 明文密钥仅此一次返回，不落库不回显
        result['plain_key'] = plain
        return Response(result, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        self._sync_grants(self.get_object(), request.data.get('grants'))
        return Response(ApiKeySerializer(self.get_object()).data)

    @action(detail=True, methods=['post'], url_path='rotate')
    def rotate(self, request, pk=None):
        """轮换密钥：生成新明文（旧密钥立即失效），授权关系不变"""
        from . import open_api_auth as auth
        api_key = self.get_object()
        if api_key.status != ApiKey.Status.ACTIVE:
            return Response({'detail': '已吊销的密钥不能轮换'}, status=status.HTTP_400_BAD_REQUEST)
        plain = auth.generate_api_key()
        api_key.key_prefix = auth.key_prefix(plain)
        api_key.key_hash = auth.hash_api_key(plain)
        api_key.save(update_fields=['key_prefix', 'key_hash'])
        result = ApiKeySerializer(api_key).data
        result['plain_key'] = plain
        return Response(result)

    @action(detail=True, methods=['post'], url_path='revoke')
    def revoke(self, request, pk=None):
        """吊销密钥：立即失效，不可恢复"""
        api_key = self.get_object()
        if api_key.status == ApiKey.Status.REVOKED:
            return Response({'detail': '该密钥已吊销'}, status=status.HTTP_400_BAD_REQUEST)
        api_key.status = ApiKey.Status.REVOKED
        api_key.revoked_at = timezone.now()
        api_key.save(update_fields=['status', 'revoked_at'])
        return Response(ApiKeySerializer(api_key).data)

    @action(detail=True, methods=['get'], url_path='call-logs')
    def call_logs(self, request, pk=None):
        """该密钥的调用日志（分页，可按 api 过滤）"""
        api_key = self.get_object()
        qs = ApiCallLog.objects.filter(api_key=api_key).select_related('api')
        api_id = request.query_params.get('api')
        if api_id:
            qs = qs.filter(api_id=api_id)
        try:
            page = max(1, int(request.query_params.get('page', 1)))
            page_size = max(1, min(int(request.query_params.get('page_size', 20)), 200))
        except (TypeError, ValueError):
            page, page_size = 1, 20
        total = qs.count()
        rows = qs[(page - 1) * page_size: page * page_size]
        return Response({
            'count': total, 'page': page, 'page_size': page_size,
            'results': ApiCallLogSerializer(rows, many=True).data,
        })


@api_view(['GET'])
def api_call_stats(request):
    """近 7 天 API 调用统计：按日趋势 + 按接口汇总（含错误数）"""
    from datetime import timedelta
    now = timezone.now()
    since = now - timedelta(days=7)
    logs = ApiCallLog.objects.filter(created_at__gte=since).values_list(
        'created_at', 'api_id', 'status_code')
    daily = {}
    per_api = {}
    total = 0
    errors = 0
    for created_at, api_id, code in logs:
        day = created_at.date().isoformat()
        daily[day] = daily.get(day, 0) + 1
        total += 1
        if code >= 400:
            errors += 1
        if api_id:
            item = per_api.setdefault(api_id, {'calls': 0, 'errors': 0})
            item['calls'] += 1
            if code >= 400:
                item['errors'] += 1
    api_names = {a.id: a.name for a in ArchiveApi.objects.filter(id__in=per_api.keys())}
    return Response({
        'total': total,
        'errors': errors,
        'daily': [{'date': d, 'calls': daily[d]} for d in sorted(daily)],
        'per_api': [{'api': aid, 'api_name': api_names.get(aid, ''), **v}
                    for aid, v in sorted(per_api.items(), key=lambda x: -x[1]['calls'])],
    })


@api_view(['GET'])
def domain_change_stats(request):
    """返回每个域的变更概况统计（档案数、最近变更时间、最近7天变更数）"""
    from datetime import timedelta
    now = timezone.now()
    seven_days_ago = now - timedelta(days=7)

    domains = Domain.objects.all().order_by('name')
    result = []
    for d in domains:
        archives = Archive.objects.filter(domain=d)
        archive_ids = list(archives.values_list('id', flat=True))
        archive_count = len(archive_ids)
        if archive_count == 0:
            continue
        last_change = ArchiveChangeDetail.objects.filter(
            archive_id__in=archive_ids
        ).aggregate(last=Max('created_at'))['last']
        change_count_7d = ArchiveChangeDetail.objects.filter(
            archive_id__in=archive_ids,
            created_at__gte=seven_days_ago
        ).count()
        result.append({
            'domain_id': d.id,
            'domain_name': d.name,
            'archive_count': archive_count,
            'last_change_at': last_change,
            'change_count_7d': change_count_7d,
        })
    return Response(result)
